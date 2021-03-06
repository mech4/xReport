import sys,os
from openpyxl import load_workbook as xloader
import xlsxwriter
import com.ihsan.lib.remotequery as rqlib
import zipfile

def runQuery(config, params, returns):
  RQ = rqlib.RQSQL(config)
  RQ.handleOperation(params, returns)

def FormOnSetDataEx(uideflist, params):
  # procedure(uideflist: TPClassUIDefList; params: TPClassUIDataPacket)
  config = uideflist.config
  if params.FirstRecord in (None,'',0):
    tbl = 'tmp_ls10'
  else:
    tbl = params.FirstRecord.tbl
  uideflist.PrepareReturnDataset()
  mlu = config.ModLibUtils
  rq = rqlib.RQSQL(config)
  rq.SELECTFROMClause = '''  
              SELECT * 
              from pbstmp.{0}
  '''.format(tbl) 
  rq.WHEREClause = '''
              1=1  
  '''
  rq.keyFieldName = 'NOMOR_REKENING'
  rq.setAltOrderFieldNames('NOMOR_REKENING')
  rq.setBaseOrderFieldNames('NOMOR_REKENING')
  rq.initOperation(uideflist.DataPacket)
  pass

def GetData(config, params, returns):
  status = returns.CreateValues(['Err', ''])
  mlu = config.ModLibUtils
  app = config.AppObject
  app.ConCreate('out')
  tempfile = config.HomeDir + 'templates\\tmp_ls10.xlsx'
  app.ConWriteln(tempfile)
  try:
    app.ConWriteln('Generating file...')
    wb = xlsxwriter.Workbook(tempfile)
    ws = wb.add_worksheet('Rekening')
    fmt2 = wb.add_format()
    fmt2.set_bold()
    fmt2.set_bg_color('yellow')
    fmt2.set_align('center')
    app.ConWriteln('get column name for rekening')
    s = '''
          select column_name
          from all_tab_columns 
          where 
          table_name = 'TMP_LS10'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    xcol = 0 
    while not res.Eof:
      ws.write(0, xcol, res.column_name, fmt2)
      ws.set_column(0, xcol, len(res.column_name)+2)
      xcol += 1
      res.Next()
    app.ConWriteln('get rekening data')
    s = '''
      select * from pbstmp.tmp_ls10
    '''
    res = config.CreateSQL(s).RawResult
    xrow = 1
    while not res.Eof:
      for xcol in range(res.FieldCount):
        ws.write(xrow, xcol, res.GetFieldValueAt(xcol))
      xrow += 1
      res.Next()
    app.ConWriteln('get column name for agunan')
    s = '''
          select column_name
          from all_tab_columns 
          where 
          table_name = 'TMP_LS10_AGUNAN'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    ws = wb.add_worksheet("Agunan")
    xcol = 0 
    while not res.Eof:
      ws.write(0, xcol, res.column_name, fmt2)
      ws.set_column(0, xcol, len(res.column_name)+6)
      xcol += 1
      res.Next()
    app.ConWriteln('get agunan data')
    s = '''
      select * from pbstmp.tmp_ls10_agunan
    '''
    res = config.CreateSQL(s).RawResult
    xrow = 1
    while not res.Eof:
      for xcol in range(res.FieldCount):
        ws.write(xrow, xcol, res.GetFieldValueAt(xcol))
      xrow += 1
      res.Next()
    #save the xlsx
    wb.close()
    sw = returns.AddStreamWrapper()
    sw.LoadFromFile(tempfile)
  except:
    status.Err = str(sys.exc_info()[1])
    app.ConRead('a')
    
def SetData(config, params, returns):
  status = returns.CreateValues(['Err', ''])
  mlu = config.ModLibUtils
  app = config.AppObject
  app.ConCreate('out')
  tempfile = config.HomeDir + 'templates\\tmp_ls10.xlsx'
  config.BeginTransaction()
  try:
    if params.StreamWrapperCount > 0:
        sw = params.GetStreamWrapper(0)
    else:
        raise Exception, 'PERINGATAN!. Download stream not found'
    sw.SaveToFile(tempfile)
    wb = xloader(tempfile)
    app.ConWriteln('Cleanup existing rekening data...')
    s = '''
          delete from pbstmp.tmp_ls10
    '''
    config.ExecSQL(s)
    app.ConWriteln('Cleanup existing agunan data...')
    s = '''
          delete from pbstmp.tmp_ls10_agunan
    '''
    config.ExecSQL(s)
    app.ConWriteln('Reading rekening data...')
    ws = wb.get_sheet_by_name(name = 'Rekening')
    app.ConWriteln('getting column info (rekening)..')
    s = '''
          select column_name, data_type
          from all_tab_columns 
          where 
          table_name = 'TMP_LS10'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    collist = {}
    colstr = '('
    while not res.Eof:
      collist[res.column_name] = res.data_type
      colstr += res.column_name + ', '
      res.Next()
    colstr=colstr.rstrip(', ') + ')'
    #app.ConWriteln(str(collist))
    xrow = 1
    testcell = ws.cell(row = xrow, column=0)
    testvalue = testcell.value
    app.ConWriteln('getting row data (rekening)..')
    while testvalue not in (None,'',0,'None'):
      xcol = 0
      valstr = '('
      for colname in collist.keys():
        getcell = ws.cell(row = xrow, column = xcol)
        getvalue = getcell.value
        if collist[colname] != 'NUMBER':
          getvalue = mlu.QuotedStr(str(getvalue))
        valstr += str(getvalue) + ', '
        xcol += 1 
      valstr=valstr.rstrip(', ') + ')'
      s = '''
        insert into pbstmp.tmp_ls10 {0} values {1} 
      '''.format(colstr, valstr)
      config.ExecSQL(s)
      xrow += 1
      testcell = ws.cell(row = xrow, column=0)
      testvalue = testcell.value
    app.ConWriteln('{0} row(s) data updated (rekening).'.format(xrow-1))
    app.ConWriteln('Reading agunan data...')
    ws = wb.get_sheet_by_name(name = 'Agunan')
    app.ConWriteln('getting column info (agunan)..')
    s = '''
          select column_name, data_type
          from all_tab_columns 
          where 
          table_name = 'TMP_LS10_AGUNAN'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    collist = {}
    colstr = '('
    while not res.Eof:
      collist[res.column_name] = res.data_type
      colstr += res.column_name + ', '
      res.Next()
    colstr=colstr.rstrip(', ') + ')'
    xrow = 1
    testcell = ws.cell(row = xrow, column=0)
    testvalue = testcell.value
    app.ConWriteln('getting row data (agunan)..')
    while testvalue not in (None,'',0,'None'):
      xcol = 0
      valstr = '('
      for colname in collist.keys():
        getcell = ws.cell(row = xrow, column = xcol)
        getvalue = getcell.value
        if collist[colname] != 'NUMBER':
          getvalue = mlu.QuotedStr(str(getvalue))
        valstr += str(getvalue) + ', '
        xcol += 1 
      valstr=valstr.rstrip(', ') + ')'
      s = '''
        insert into pbstmp.tmp_ls10_agunan {0} values {1} 
      '''.format(colstr, valstr)
      config.ExecSQL(s)
      xrow += 1
      testcell = ws.cell(row = xrow, column=0)
      testvalue = testcell.value
    app.ConWriteln('{0} row(s) data updated (agunan).'.format(xrow-1))
    config.Commit()
  except:
    config.Rollback()
    status.Err = str(sys.exc_info()[1])
    app.ConRead('a')

def GetCSVData(config, params, returns):
  status = returns.CreateValues(['Err', ''])
  mlu = config.ModLibUtils
  app = config.AppObject
  app.ConCreate('out')
  rtmp = config.HomeDir + 'templates\\rekening_ls10.csv'
  rname = 'rekening_ls10.csv'
  atmp = config.HomeDir + 'templates\\agunan_ls10.csv'
  aname = 'agunan_ls10.csv'
  ztmp = config.HomeDir + 'templates\\tmp_ls10.zip'
  tmploc = config.HomeDir + 'templates' 
  try:
    app.ConWriteln('Generating file...')
    rkfile = open(rtmp, 'w')
    app.ConWriteln('get column name for rekening')
    s = '''
          select column_name
          from all_tab_columns 
          where 
          table_name = 'TMP_LS10'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    rowheader = ''
    while not res.Eof:
      rowheader += res.column_name
      rowheader += ';'
      res.Next()
    rkfile.write(rowheader.rstrip(';') + '\n')
    app.ConWriteln('get rekening data')
    s = '''
      select * from pbstmp.tmp_ls10
    '''
    res = config.CreateSQL(s).RawResult
    while not res.Eof:
      rowcontent = '' 
      for xcol in range(res.FieldCount):
        rowcontent += str(res.GetFieldValueAt(xcol))
        rowcontent += ';'
      rkfile.write(rowcontent.rstrip(';')+'\n')
      res.Next()
    rkfile.close()
    app.ConWriteln('get column name for agunan')
    s = '''
          select column_name
          from all_tab_columns 
          where 
          table_name = 'TMP_LS10_AGUNAN'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    agufile = open(atmp, 'w')
    rowheader = ''
    while not res.Eof:
      rowheader += res.column_name
      rowheader += ';'
      res.Next()
    agufile.write(rowheader.rstrip(';')+'\n')
    app.ConWriteln('get agunan data')
    s = '''
      select * from pbstmp.tmp_ls10_agunan
    '''
    res = config.CreateSQL(s).RawResult
    while not res.Eof:
      rowcontent = ''
      for xcol in range(res.FieldCount):
        rowcontent += str(res.GetFieldValueAt(xcol))
        rowcontent += ';'
      agufile.write(rowcontent.rstrip(';')+'\n')
      res.Next()
    agufile.close()
    zf = zipfile.ZipFile(ztmp, mode='w')
    zf.write(rtmp, rname)
    zf.write(atmp, aname)
    zf.close()
    sw = returns.AddStreamWrapper()
    sw.LoadFromFile(ztmp)
  except:
    status.Err = str(sys.exc_info()[1])
    app.ConWriteln(status.Err)
    app.ConRead('a')
    
def SetCSVData(config, params, returns):
  status = returns.CreateValues(['Err', ''])
  mlu = config.ModLibUtils
  app = config.AppObject
  app.ConCreate('out')
  rtmp = config.HomeDir + 'templates\\rekening_ls10.csv'
  rname = 'rekening_ls10.csv'
  atmp = config.HomeDir + 'templates\\agunan_ls10.csv'
  aname = 'agunan_ls10.csv'
  ztmp = config.HomeDir + 'templates\\tmp_ls10.zip'
  tmploc = config.HomeDir + 'templates' 
  config.BeginTransaction()
  try:
    if params.StreamWrapperCount > 0:
        sw = params.GetStreamWrapper(0)
    else:
        raise Exception, 'PERINGATAN!. Download stream not found'
    sw.SaveToFile(ztmp)
    zf = zipfile.ZipFile(ztmp)
    zf.extract(rname, tmploc)
    zf.extract(aname, tmploc)
    zf.close()
    app.ConWriteln('Cleanup existing rekening data...')
    s = '''
          delete from pbstmp.tmp_ls10
    '''
    config.ExecSQL(s)
    app.ConWriteln('Cleanup existing agunan data...')
    s = '''
          delete from pbstmp.tmp_ls10_agunan
    '''
    config.ExecSQL(s)
    app.ConWriteln('Reading rekening data...')
    rkfile = open(rtmp)
    app.ConWriteln('getting column info (rekening)..')
    s = '''
          select column_name, data_type
          from all_tab_columns 
          where 
          table_name = 'TMP_LS10'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    collist = {}
    colstr = '('
    while not res.Eof:
      collist[res.column_name] = res.data_type
      colstr += res.column_name + ', '
      res.Next()
    colstr=colstr.rstrip(', ') + ')'
    #app.ConWriteln(str(collist))
    rkfile.readline()
    testvalue = rkfile.readline()
    xrow = 0
    app.ConWriteln('getting row data (rekening)..')
    while testvalue not in (None,'',0,'None'):
      xcol = 0
      vallist = testvalue.rstrip('\n').split(';')
      valstr = '('
      for colname in collist.keys():
        getvalue = vallist[xcol]
        if collist[colname] != 'NUMBER':
          getvalue = mlu.QuotedStr(str(getvalue))
        valstr += str(getvalue) + ', '
        xcol += 1 
      valstr=valstr.rstrip(', ') + ')'
      s = '''
        insert into pbstmp.tmp_ls10 {0} values {1} 
      '''.format(colstr, valstr)
      config.ExecSQL(s)
      xrow += 1
      testvalue = rkfile.readline()
    app.ConWriteln('{0} row(s) data updated (rekening).'.format(xrow))
    rkfile.close()
    app.ConWriteln('Reading agunan data...')
    agufile = open(atmp)
    app.ConWriteln('getting column info (agunan)..')
    s = '''
          select column_name, data_type
          from all_tab_columns 
          where 
          table_name = 'TMP_LS10_AGUNAN'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    collist = {}
    colstr = '('
    while not res.Eof:
      collist[res.column_name] = res.data_type
      colstr += res.column_name + ', '
      res.Next()
    colstr=colstr.rstrip(', ') + ')'
    xrow = 0
    agufile.readline()
    testvalue = agufile.readline()
    app.ConWriteln('getting row data (agunan)..')
    while testvalue not in (None,'',0,'None'):
      xcol = 0
      vallist = testvalue.rstrip('\n').split(';')
      valstr = '('
      for colname in collist.keys():
        getvalue = vallist[xcol]
        if collist[colname] != 'NUMBER':
          getvalue = mlu.QuotedStr(str(getvalue))
        valstr += str(getvalue) + ', '
        xcol += 1 
      valstr=valstr.rstrip(', ') + ')'
      s = '''
        insert into pbstmp.tmp_ls10_agunan {0} values {1} 
      '''.format(colstr, valstr)
      config.ExecSQL(s)
      xrow += 1
      testvalue = agufile.readline()
    app.ConWriteln('{0} row(s) data updated (agunan).'.format(xrow))
    agufile.close()
    config.Commit()
  except:
    config.Rollback()
    status.Err = str(sys.exc_info()[1])
    app.ConRead('a')
    
def GenereateInstance(config, params, returns):
  app = config.AppObject

  SQL = '''
      select
        to_char(NOMOR_NASABAH) 	          as NOMOR_NASABAH,
        to_char(GOLONGAN_DEBITUR) 	          as GOLONGAN_DEBITUR,
        to_char(HUB_DGN_BANK) 	          as HUB_DGN_BANK,
        to_char(NASABAH_STATUS) 	          as NASABAH_STATUS,
        to_char(KATEGORI_USAHA) 	          as KATEGORI_USAHA,
        to_char(NOMOR_REKENING) 	          as NOMOR_REKENING,
        to_char(JENIS_VALUTA) 	          as JENIS_VALUTA,
        to_char(BLNTHN_MULAI) 	          as BLNTHN_MULAI,
        to_char(BLNTHN_JATUH_TEMPO) 	          as BLNTHN_JATUH_TEMPO,
        to_char(SUMBER_DANA) 	          as SUMBER_DANA,
        to_char(LOKASI_PROYEK) 	          as LOKASI_PROYEK,
        to_char(JENIS_PIUTANG) 	          as JENIS_PIUTANG,
        to_char(SIFAT_PIUTANG) 	          as SIFAT_PIUTANG,
        to_char(JENIS_PENGGUNAAN) 	          as JENIS_PENGGUNAAN,
        to_char(ORIENTASI_PENGGUNANAN) 	          as ORIENTASI_PENGGUNANAN,
        to_char(SEKTOR_EKONOMI) 	          as SEKTOR_EKONOMI,
        to_char(KATEGORI_PORTOFOLIO) 	          as KATEGORI_PORTOFOLIO,
        to_char(LEMBAGA_PEMERINGKAT) 	          as LEMBAGA_PEMERINGKAT,
        to_char(NILAI_PERINGKAT) 	          as NILAI_PERINGKAT,
        to_char(TGL_PERINGKAT) 	          as TGL_PERINGKAT,
        to_char(NILAI_KONTRAK) 	          as NILAI_KONTRAK,
        to_char(PERSEN_AWAL) 	          as PERSEN_AWAL,
        to_char(PERSEN_AKHIR) 	          as PERSEN_AKHIR,
        to_char(PIUTANG) 	          as PIUTANG,
        to_char(POKOK) 	          as POKOK,
        to_char(MARGIN) 	          as MARGIN,
        to_char(BAKI_LALU) 	          as BAKI_LALU,
        to_char(BAKI_LAPOR) 	          as BAKI_LAPOR,
        to_char(TUNGGAKAN_HARI) 	          as TUNGGAKAN_HARI,
        to_char(TUNGGAKAN_POKOK) 	          as TUNGGAKAN_POKOK,
        to_char(TUNGGAKAN_MARGIN) 	          as TUNGGAKAN_MARGIN,
        to_char(IMBALAN_DITERIMA) 	          as IMBALAN_DITERIMA,
        to_char(KOLEKTIBILITAS) 	          as KOLEKTIBILITAS,
        to_char(CKPN_INDIVIDUAL) 	          as CKPN_INDIVIDUAL,
        to_char(CKPN_KOLEKTIF) 	          as CKPN_KOLEKTIF
        from {1}
  '''.format(
    config.MapDBTableName('financing.finaccount'), # 0
    config.MapDBTableName('tmp.TMP_LS10')                     
  )
  
  rSQL =config.CreateSQL(SQL).RawResult   
  
  ResultDir = 'c:/dafapp/rawdata/'

  sBaseFileName = '517001000-2013-12-31-MM-BSMS56.xml'
  sFileName = ResultDir + sBaseFileName
  
  i = 0
  txt_header ="""<?xml version="1.0" encoding="UTF-8"?>
<xbrli:xbrl xmlns:base="http://xbrl.bi.go.id/xbrl/2013-11-01/dict/base" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:t="http://xbrl.bi.go.id/xbrl/2013-11-01/dict/types" xmlns:BSMS8="http://xbrl.bi.go.id/xbrl/2013-11-01/view/Pelaporan_Keuangan/Rincian_rupa_rupa_aset_dan_kewajiban/BSMS8" xmlns:BSMS56="http://xbrl.bi.go.id/xbrl/2013-11-01/view/Pelaporan_Keuangan/Rincian_kredit/BSMS56" xmlns:BSMS20="http://xbrl.bi.go.id/xbrl/2013-11-01/view/Pelaporan_Keuangan/Rincian_rupa_rupa_aset_dan_kewajiban/BSMS20" xmlns:xlink="http://www.w3.org/1999/xlink" xmlns:xl="http://www.xbrl.org/2003/XLink" xmlns:link="http://www.xbrl.org/2003/linkbase" xmlns:bf="http://xbrl.org/2008/filter/boolean" xmlns:gen="http://xbrl.org/2008/generic" xmlns:label="http://xbrl.org/2008/label" xmlns:variable="http://xbrl.org/2008/variable" xmlns:msg="http://xbrl.org/2010/message" xmlns:valm="http://xbrl.org/2010/message/validation" xmlns:xbrli="http://www.xbrl.org/2003/instance">
<link:schemaRef xlink:type="simple" xlink:href="http://xbrl.bi.go.id/xbrl/2013-11-01/view/Pelaporan%20Keuangan/Rincian%20kredit/BSMS56/BSMS56.xsd"/>
<xbrli:context id="c1">
<xbrli:entity>
<xbrli:identifier scheme="PBS">517001000</xbrli:identifier>
</xbrli:entity>
<xbrli:period>
<xbrli:instant>2013-11-30</xbrli:instant>
</xbrli:period>
</xbrli:context>
<base:dummy contextRef="c1" unitRef="I" xsi:nil="true"/>\n"""
  
  txt_form1 = """<BSMS56:BSMS56 id="ID_{0}">
<base:si68 contextRef="c1">{1}</base:si68>
<base:si421 contextRef="c1">{2}</base:si421>
<base:si15 contextRef="c1">{3}</base:si15>
<base:si16 contextRef="c1">{4}</base:si16>
<base:si83 contextRef="c1">{5}</base:si83>
<base:si188 contextRef="c1">{6}</base:si188>
<base:si2 contextRef="c1">{7}</base:si2>
<base:di9 contextRef="c1">{8}</base:di9>
<base:di10 contextRef="c1">{9}</base:di10>
<base:si11 contextRef="c1">{10}</base:si11>
<base:si84 contextRef="c1">{11}</base:si84>
<base:si85 contextRef="c1">{12}</base:si85>
<base:si86 contextRef="c1">{13}</base:si86>
<base:si434 contextRef="c1">{14}</base:si434>
<base:si87 contextRef="c1">{15}</base:si87>
<base:si394 contextRef="c1">{16}</base:si394>
<base:si140 contextRef="c1">{17}</base:si140>
<base:si71 contextRef="c1">{18}</base:si71>
<base:si19 contextRef="c1">{19}</base:si19>
<base:di72 contextRef="c1" xsi:nil="true" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"/>
<base:mi88 contextRef="c1" unitRef="I" decimals="2">{21}</base:mi88>
<base:pi50 contextRef="c1" unitRef="I" decimals="2">{22}</base:pi50>
<base:pi12 contextRef="c1" unitRef="I" decimals="2">{23}</base:pi12>
<base:mi6016 contextRef="c1" unitRef="I" decimals="2">{24}</base:mi6016>
<base:mi89 contextRef="c1" unitRef="I" decimals="2">{25}</base:mi89>
<base:mi90 contextRef="c1" unitRef="I" decimals="2">{26}</base:mi90>
<base:mi91 contextRef="c1" unitRef="I" decimals="2">{27}</base:mi91>
<base:mi92 contextRef="c1" unitRef="I" decimals="2">{28}</base:mi92>
<base:mi93 contextRef="c1" unitRef="I" decimals="2">{29}</base:mi93>
<base:mi94 contextRef="c1" unitRef="I" decimals="2">{30}</base:mi94>
<base:mi95 contextRef="c1" unitRef="I" decimals="2">{31}</base:mi95>
<base:mi96 contextRef="c1" unitRef="I" decimals="2">{32}</base:mi96>
<base:si58 contextRef="c1">{33}</base:si58>
"""

  txt_form2 = """\n
<BSMS56:BSMS56-a>
<base:si27 contextRef="c1">{0}</base:si27>
<base:si28 contextRef="c1">{1}</base:si28>
<base:si80 contextRef="c1">{2}</base:si80>
<base:si29 contextRef="c1">{3}</base:si29>
<base:di30 contextRef="c1">{4}</base:di30>
<base:mi31 contextRef="c1" unitRef="I" decimals="2">{5}</base:mi31>
<base:mi60 contextRef="c1" unitRef="I" decimals="2">{6}</base:mi60>
<base:pi59 contextRef="c1" unitRef="I" decimals="2">{7}</base:pi59>
</BSMS56:BSMS56-a>
"""

  txt_form3 = """\n
<base:mi97 contextRef="c1" unitRef="I" decimals="2">{0}</base:mi97>
<base:mi98 contextRef="c1" unitRef="I" decimals="2">{1}</base:mi98>
</BSMS56:BSMS56>"""

  txt_footer ="""\n
<xbrli:unit id="I">
<xbrli:measure xmlns:iso4217="http://www.xbrl.org/2003/iso4217">iso4217:IDR</xbrli:measure>
</xbrli:unit>
</xbrli:xbrl>"""
  
  oFile = open(sFileName,'w')
  try:
    oFile.write(txt_header)
    _startLine = 1 
    while not rSQL.Eof:
      i = i+1 
      oFile.write(txt_form1.format(
        _startLine, #0
        rSQL.NOMOR_NASABAH, # 1
        rSQL.GOLONGAN_DEBITUR, # 1
        rSQL.HUB_DGN_BANK, # 1
        rSQL.NASABAH_STATUS, # 1
        rSQL.KATEGORI_USAHA, # 1
        rSQL.NOMOR_REKENING, # 1
        rSQL.JENIS_VALUTA, # 1
        rSQL.BLNTHN_MULAI, # 1
        rSQL.BLNTHN_JATUH_TEMPO, # 1
        rSQL.SUMBER_DANA, # 1
        rSQL.LOKASI_PROYEK, # 1
        rSQL.JENIS_PIUTANG, # 1
        rSQL.SIFAT_PIUTANG, # 1
        rSQL.JENIS_PENGGUNAAN, # 1
        rSQL.ORIENTASI_PENGGUNANAN, # 1
        rSQL.SEKTOR_EKONOMI, # 1
        rSQL.KATEGORI_PORTOFOLIO, # 1
        rSQL.LEMBAGA_PEMERINGKAT, # 1
        rSQL.NILAI_PERINGKAT, # 1
        rSQL.TGL_PERINGKAT, # 1
        rSQL.NILAI_KONTRAK, # 1
        rSQL.PERSEN_AWAL, # 1
        rSQL.PERSEN_AKHIR, # 1
        rSQL.PIUTANG, # 1
        rSQL.POKOK, # 1
        rSQL.MARGIN, # 1
        rSQL.BAKI_LALU, # 1
        rSQL.BAKI_LAPOR, # 1
        rSQL.TUNGGAKAN_HARI, # 1
        rSQL.TUNGGAKAN_POKOK, # 1
        rSQL.TUNGGAKAN_MARGIN, # 1
        rSQL.IMBALAN_DITERIMA, # 1
        rSQL.KOLEKTIBILITAS, # 1
        '' # 
      ))        
      SQL_agunan = '''
      select
        to_char(JENIS_ASSET) 	          as JENIS_ASSET,
        to_char(NO_AGUNAN) 	          as NO_AGUNAN,
        to_char(SIFAT_AGUNAN) 	          as SIFAT_AGUNAN,
        to_char(GOLONGAN_PENERBIT) 	          as GOLONGAN_PENERBIT,
        to_char(TGL_PENILAIAN) 	          as TGL_PENILAIAN,
        to_char(NILAI_AGUNAN) 	          as NILAI_AGUNAN
        from {0} where nomor_rekening = '{1}'
        '''.format(
          config.MapDBTableName('tmp.TMP_LS10_agunan'), #0 
          rSQL.NOMOR_REKENING # 0
        )
      rSQL_a =config.CreateSQL(SQL_agunan).RawResult  
      while not rSQL_a.Eof:
      
          oFile.write(txt_form2.format(
            rSQL_a.JENIS_ASSET, # 0
            rSQL_a.NO_AGUNAN, # 1
            rSQL_a.SIFAT_AGUNAN, # 2
            rSQL_a.GOLONGAN_PENERBIT, # 3
            rSQL_a.TGL_PENILAIAN, # 4
            rSQL_a.NILAI_AGUNAN, # 5
            '',#rSQL_a.DPT_DIPERHITUNGKAN, # 6
            '',#rSQL_a.BAGIAN_DIJAMINKAN, # 7      
            '' # 
          ))
          rSQL_a.Next()
      
      oFile.write(txt_form3.format(
        rSQL.CKPN_INDIVIDUAL, # 0
        rSQL.CKPN_KOLEKTIF, # 1
      ))
      
      if (i%100 == 0):
        printOut("Proses data ke-%d " % i)

      _startLine=_startLine+1
      rSQL.Next()
    #--

    
    if ResultDir == '' :                                
      ResultDir = 'c:/dafapp/rawdata/'

    oFile.write(txt_footer)
  

  finally:
    oFile.close()
    ztmp = sFileName.replace('.xml','.zip')
    zf = zipfile.ZipFile(ztmp, mode='w')
    zf.write(sFileName, sBaseFileName,zipfile.ZIP_DEFLATED)
    zf.close()
    os.remove(sFileName)
    sw = returns.AddStreamWrapper()
    sw.LoadFromFile(ztmp)
    sw.MIMEType = 'zip'
    
def GenereateCtl(config, params, returns):
  app = config.AppObject

  SQL = '''
      select
        to_char(NOMOR_NASABAH) 	          as NOMOR_NASABAH,
        to_char(GOLONGAN_DEBITUR) 	          as GOLONGAN_DEBITUR,
        to_char(HUB_DGN_BANK) 	          as HUB_DGN_BANK,
        to_char(NASABAH_STATUS) 	          as NASABAH_STATUS,
        to_char(KATEGORI_USAHA) 	          as KATEGORI_USAHA,
        to_char(NOMOR_REKENING) 	          as NOMOR_REKENING,
        to_char(JENIS_VALUTA) 	          as JENIS_VALUTA,
        to_char(BLNTHN_MULAI) 	          as BLNTHN_MULAI,
        to_char(BLNTHN_JATUH_TEMPO) 	          as BLNTHN_JATUH_TEMPO,
        to_char(SUMBER_DANA) 	          as SUMBER_DANA,
        to_char(LOKASI_PROYEK) 	          as LOKASI_PROYEK,
        to_char(JENIS_PIUTANG) 	          as JENIS_PIUTANG,
        to_char(SIFAT_PIUTANG) 	          as SIFAT_PIUTANG,
        to_char(JENIS_PENGGUNAAN) 	          as JENIS_PENGGUNAAN,
        to_char(ORIENTASI_PENGGUNANAN) 	          as ORIENTASI_PENGGUNANAN,
        to_char(SEKTOR_EKONOMI) 	          as SEKTOR_EKONOMI,
        to_char(KATEGORI_PORTOFOLIO) 	          as KATEGORI_PORTOFOLIO,
        to_char(LEMBAGA_PEMERINGKAT) 	          as LEMBAGA_PEMERINGKAT,
        to_char(NILAI_PERINGKAT) 	          as NILAI_PERINGKAT,
        to_char(TGL_PERINGKAT) 	          as TGL_PERINGKAT,
        to_char(NILAI_KONTRAK) 	          as NILAI_KONTRAK,
        to_char(PERSEN_AWAL) 	          as PERSEN_AWAL,
        to_char(PERSEN_AKHIR) 	          as PERSEN_AKHIR,
        to_char(PIUTANG) 	          as PIUTANG,
        to_char(POKOK) 	          as POKOK,
        to_char(MARGIN) 	          as MARGIN,
        to_char(BAKI_LALU) 	          as BAKI_LALU,
        to_char(BAKI_LAPOR) 	          as BAKI_LAPOR,
        to_char(TUNGGAKAN_HARI) 	          as TUNGGAKAN_HARI,
        to_char(TUNGGAKAN_POKOK) 	          as TUNGGAKAN_POKOK,
        to_char(TUNGGAKAN_MARGIN) 	          as TUNGGAKAN_MARGIN,
        to_char(IMBALAN_DITERIMA) 	          as IMBALAN_DITERIMA,
        to_char(KOLEKTIBILITAS) 	          as KOLEKTIBILITAS,
        to_char(CKPN_INDIVIDUAL) 	          as CKPN_INDIVIDUAL,
        to_char(CKPN_KOLEKTIF) 	          as CKPN_KOLEKTIF
        from {1}
  '''.format(
    config.MapDBTableName('financing.finaccount'), # 0
    config.MapDBTableName('tmp.TMP_LS10')                     
  )
  
  rSQL =config.CreateSQL(SQL).RawResult   
  
  ResultDir = 'c:/dafapp/rawdata/'

  sBaseFileName = 'BSMS56.ctl'
  sFileName = ResultDir + sBaseFileName
  
  i = 0
  txt_header ="""LOAD DATA
INFILE *
TRUNCATE
INTO TABLE pbstmp.tmp_ls10
FIELDS TERMINATED BY '|'
TRAILING NULLCOLS ( 
NOMOR_NASABAH,
GOLONGAN_DEBITUR,
HUB_DGN_BANK,
NASABAH_STATUS,
KATEGORI_USAHA,
NOMOR_REKENING,
JENIS_VALUTA,
BLNTHN_MULAI,
BLNTHN_JATUH_TEMPO,
SUMBER_DANA,
LOKASI_PROYEK,
JENIS_PIUTANG,
SIFAT_PIUTANG,
JENIS_PENGGUNAAN,
ORIENTASI_PENGGUNANAN,
SEKTOR_EKONOMI,
KATEGORI_PORTOFOLIO,
LEMBAGA_PEMERINGKAT,
NILAI_PERINGKAT,
TGL_PERINGKAT,
NILAI_KONTRAK,
PERSEN_AWAL,
PERSEN_AKHIR,
PIUTANG,
POKOK,
MARGIN,
BAKI_LALU,
BAKI_LAPOR,
TUNGGAKAN_HARI,
TUNGGAKAN_POKOK,
TUNGGAKAN_MARGIN,
IMBALAN_DITERIMA,
KOLEKTIBILITAS,
DPT_DIPERHITUNGKAN,
BAGIAN_DIJAMINKAN,
CKPN_INDIVIDUAL,
CKPN_KOLEKTIF)
BEGINDATA\n"""
  
  txt_form1 = """{1}|{2}|{3}|{4}|{5}|{6}|{7}|{8}|{9}|{10}|{11}|{12}|{13}|{14}|{15}|{16}|{17}|{18}|{19}|{20}|{21}|{22}|{23}|{24}|{25}|{26}|{27}|{28}|{29}|{30}|{31}|{32}|{33}|{34}|{35}
"""

  oFile = open(sFileName,'w')
  try:
    oFile.write(txt_header)
    _startLine = 1 
    while not rSQL.Eof:
      i = i+1 
      oFile.write(txt_form1.format(
        _startLine, #0
        rSQL.NOMOR_NASABAH, # 1
        rSQL.GOLONGAN_DEBITUR, # 1
        rSQL.HUB_DGN_BANK, # 1
        rSQL.NASABAH_STATUS, # 1
        rSQL.KATEGORI_USAHA, # 1
        rSQL.NOMOR_REKENING, # 1
        rSQL.JENIS_VALUTA, # 1
        rSQL.BLNTHN_MULAI, # 1
        rSQL.BLNTHN_JATUH_TEMPO, # 1
        rSQL.SUMBER_DANA, # 1
        rSQL.LOKASI_PROYEK, # 1
        rSQL.JENIS_PIUTANG, # 1
        rSQL.SIFAT_PIUTANG, # 1
        rSQL.JENIS_PENGGUNAAN, # 1
        rSQL.ORIENTASI_PENGGUNANAN, # 1
        rSQL.SEKTOR_EKONOMI, # 1
        rSQL.KATEGORI_PORTOFOLIO, # 1
        rSQL.LEMBAGA_PEMERINGKAT, # 1
        rSQL.NILAI_PERINGKAT, # 1
        rSQL.TGL_PERINGKAT, # 1
        rSQL.NILAI_KONTRAK, # 1
        rSQL.PERSEN_AWAL, # 1
        rSQL.PERSEN_AKHIR, # 1
        rSQL.PIUTANG, # 1
        rSQL.POKOK, # 1
        rSQL.MARGIN, # 1
        rSQL.BAKI_LALU, # 1
        rSQL.BAKI_LAPOR, # 1
        rSQL.TUNGGAKAN_HARI, # 1
        rSQL.TUNGGAKAN_POKOK, # 1
        rSQL.TUNGGAKAN_MARGIN, # 1
        rSQL.IMBALAN_DITERIMA, # 1
        rSQL.KOLEKTIBILITAS, # 1
        '', # 
        '', # 
        rSQL.CKPN_INDIVIDUAL, # 0
        rSQL.CKPN_KOLEKTIF, # 1
        '' # 
      ))        

      if (i%100 == 0):
        printOut("Proses data ke-%d " % i)

      _startLine=_startLine+1
      rSQL.Next()
    #--

    
    if ResultDir == '' :                                
      ResultDir = 'c:/dafapp/rawdata/'

  finally:
    oFile.close()
    ztmp = sFileName.replace('.ctl','.zip')
    zf = zipfile.ZipFile(ztmp, mode='w')
    zf.write(sFileName, sBaseFileName,zipfile.ZIP_DEFLATED)
    zf.close()
    os.remove(sFileName)
    sw = returns.AddStreamWrapper()
    sw.LoadFromFile(ztmp)
    sw.MIMEType = 'zip'
    
def GenereateCtl2(config, params, returns):
  app = config.AppObject

  SQL = '''
      select
        to_char(NOMOR_NASABAH) 	          as NOMOR_NASABAH,
        to_char(GOLONGAN_DEBITUR) 	          as GOLONGAN_DEBITUR,
        to_char(HUB_DGN_BANK) 	          as HUB_DGN_BANK,
        to_char(NASABAH_STATUS) 	          as NASABAH_STATUS,
        to_char(KATEGORI_USAHA) 	          as KATEGORI_USAHA,
        to_char(NOMOR_REKENING) 	          as NOMOR_REKENING,
        to_char(JENIS_VALUTA) 	          as JENIS_VALUTA,
        to_char(BLNTHN_MULAI) 	          as BLNTHN_MULAI,
        to_char(BLNTHN_JATUH_TEMPO) 	          as BLNTHN_JATUH_TEMPO,
        to_char(SUMBER_DANA) 	          as SUMBER_DANA,
        to_char(LOKASI_PROYEK) 	          as LOKASI_PROYEK,
        to_char(JENIS_PIUTANG) 	          as JENIS_PIUTANG,
        to_char(SIFAT_PIUTANG) 	          as SIFAT_PIUTANG,
        to_char(JENIS_PENGGUNAAN) 	          as JENIS_PENGGUNAAN,
        to_char(ORIENTASI_PENGGUNANAN) 	          as ORIENTASI_PENGGUNANAN,
        to_char(SEKTOR_EKONOMI) 	          as SEKTOR_EKONOMI,
        to_char(KATEGORI_PORTOFOLIO) 	          as KATEGORI_PORTOFOLIO,
        to_char(LEMBAGA_PEMERINGKAT) 	          as LEMBAGA_PEMERINGKAT,
        to_char(NILAI_PERINGKAT) 	          as NILAI_PERINGKAT,
        to_char(TGL_PERINGKAT) 	          as TGL_PERINGKAT,
        to_char(NILAI_KONTRAK) 	          as NILAI_KONTRAK,
        to_char(PERSEN_AWAL) 	          as PERSEN_AWAL,
        to_char(PERSEN_AKHIR) 	          as PERSEN_AKHIR,
        to_char(PIUTANG) 	          as PIUTANG,
        to_char(POKOK) 	          as POKOK,
        to_char(MARGIN) 	          as MARGIN,
        to_char(BAKI_LALU) 	          as BAKI_LALU,
        to_char(BAKI_LAPOR) 	          as BAKI_LAPOR,
        to_char(TUNGGAKAN_HARI) 	          as TUNGGAKAN_HARI,
        to_char(TUNGGAKAN_POKOK) 	          as TUNGGAKAN_POKOK,
        to_char(TUNGGAKAN_MARGIN) 	          as TUNGGAKAN_MARGIN,
        to_char(IMBALAN_DITERIMA) 	          as IMBALAN_DITERIMA,
        to_char(KOLEKTIBILITAS) 	          as KOLEKTIBILITAS,
        to_char(CKPN_INDIVIDUAL) 	          as CKPN_INDIVIDUAL,
        to_char(CKPN_KOLEKTIF) 	          as CKPN_KOLEKTIF
        from {1}
  '''.format(
    config.MapDBTableName('financing.finaccount'), # 0
    config.MapDBTableName('tmp.TMP_LS10_AGUNAN')                     
  )
  
  rSQL =config.CreateSQL(SQL).RawResult   
  
  ResultDir = 'c:/dafapp/rawdata/'

  sBaseFileName = 'BSMS56_a.ctl'
  sFileName = ResultDir + sBaseFileName
  
  i = 0
  txt_header ="""LOAD DATA
INFILE *
TRUNCATE
INTO TABLE pbstmp.tmp_ls10_agunan
FIELDS TERMINATED BY '|'
TRAILING NULLCOLS ( 
NOMOR_REKENING,
JENIS_ASSET,
NO_AGUNAN,
SIFAT_AGUNAN,
GOLONGAN_PENERBIT,
TGL_PENILAIAN,
NILAI_AGUNAN)
BEGINDATA\n"""
  
  txt_form1 = """{1}|{2}|{3}|{4}|{5}|{6}|{7}
"""

  oFile = open(sFileName,'w')
  try:
    oFile.write(txt_header)
    _startLine = 1 
    while not rSQL.Eof:
      i = i+1 
      oFile.write(txt_form1.format(
        _startLine, #0
        rSQL.NOMOR_REKENING,
        rSQL.JENIS_ASSET,
        rSQL.NO_AGUNAN,
        rSQL.SIFAT_AGUNAN,
        rSQL.GOLONGAN_PENERBIT,
        rSQL.TGL_PENILAIAN,
        rSQL.NILAI_AGUNAN
      ))        

      if (i%100 == 0):
        printOut("Proses data ke-%d " % i)

      _startLine=_startLine+1
      rSQL.Next()
    #--

    
    if ResultDir == '' :                                
      ResultDir = 'c:/dafapp/rawdata/'

  finally:
    oFile.close()
    ztmp = sFileName.replace('.ctl','.zip')
    zf = zipfile.ZipFile(ztmp, mode='w')
    zf.write(sFileName, sBaseFileName,zipfile.ZIP_DEFLATED)
    zf.close()
    os.remove(sFileName)
    sw = returns.AddStreamWrapper()
    sw.LoadFromFile(ztmp)
    sw.MIMEType = 'zip'
    
def SetCTLData1(config, params, returns):
  status = returns.CreateValues(['Err', ''])
  mlu = config.ModLibUtils
  app = config.AppObject
  app.ConCreate('out')
  rtmp = config.HomeDir + 'templates\\BSMS56.ctl'
  rname = 'BSMS56.ctl'
  ztmp = config.HomeDir + 'templates\\BSMS56.zip'
  tmploc = config.HomeDir + 'templates' 
  config.BeginTransaction()
  try:
    if params.StreamWrapperCount > 0:
        sw = params.GetStreamWrapper(0)
    else:
        raise Exception, 'PERINGATAN!. Download stream not found'
    sw.SaveToFile(ztmp)
    zf = zipfile.ZipFile(ztmp)
    zf.extract(rname, tmploc)
    zf.close()
    #app.ConWriteln('Cleanup existing rekening data...')
    #s = '''
    #      delete from pbstmp.tmp_ls10
    #'''
    #config.ExecSQL(s)
    config.Commit()
  except:
    config.Rollback()
    status.Err = str(sys.exc_info()[1])
    app.ConRead('a')
    
    



