import sys,os
from openpyxl import load_workbook as xloader
import xlsxwriter
import com.ihsan.lib.remotequery as rqlib
import zipfile

def runQuery(config, params, returns):
  RQ = rqlib.RQSQL(config)
  RQ.handleOperation(params, returns)

def FormOnSetDataEx(uideflist, params):
  # procedure(uideflist: TPClassUIDefList; params: TPClassUIDataPacket)
  config = uideflist.config
  if params.FirstRecord in (None,'',0):
    tbl = 'tmp_ls13'
  else:
    tbl = params.FirstRecord.tbl
  uideflist.PrepareReturnDataset()
  mlu = config.ModLibUtils
  rq = rqlib.RQSQL(config)
  rq.SELECTFROMClause = '''  
              SELECT * 
              from pbstmp.{0}
  '''.format(tbl) 
  rq.WHEREClause = '''
              1=1  
  '''
  rq.keyFieldName = 'NOMOR_REKENING'
  rq.setAltOrderFieldNames('NOMOR_REKENING')
  rq.setBaseOrderFieldNames('NOMOR_REKENING')
  rq.initOperation(uideflist.DataPacket)
  pass

def GetData(config, params, returns):
  status = returns.CreateValues(['Err', ''])
  mlu = config.ModLibUtils
  app = config.AppObject
  app.ConCreate('out')
  tempfile = config.HomeDir + 'templates\\tmp_ls13.xlsx'
  app.ConWriteln(tempfile)
  try:
    app.ConWriteln('Generating file...')
    wb = xlsxwriter.Workbook(tempfile)
    ws = wb.add_worksheet('Rekening')
    fmt2 = wb.add_format()
    fmt2.set_bold()
    fmt2.set_bg_color('yellow')
    fmt2.set_align('center')
    app.ConWriteln('get column name for rekening')
    s = '''
          select column_name
          from all_tab_columns 
          where 
          table_name = 'TMP_LS13'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    xcol = 0 
    while not res.Eof:
      ws.write(0, xcol, res.column_name, fmt2)
      ws.set_column(0, xcol, len(res.column_name)+2)
      xcol += 1
      res.Next()
    app.ConWriteln('get rekening data')
    s = '''
      select * from pbstmp.tmp_ls13
    '''
    res = config.CreateSQL(s).RawResult
    xrow = 1
    while not res.Eof:
      for xcol in range(res.FieldCount):
        ws.write(xrow, xcol, res.GetFieldValueAt(xcol))
      xrow += 1
      res.Next()
    app.ConWriteln('get column name for agunan')
    s = '''
          select column_name
          from all_tab_columns 
          where 
          table_name = 'TMP_LS13_AGUNAN'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    ws = wb.add_worksheet("Agunan")
    xcol = 0 
    while not res.Eof:
      ws.write(0, xcol, res.column_name, fmt2)
      ws.set_column(0, xcol, len(res.column_name)+6)
      xcol += 1
      res.Next()
    app.ConWriteln('get agunan data')
    s = '''
      select * from pbstmp.tmp_ls13_agunan
    '''
    res = config.CreateSQL(s).RawResult
    xrow = 1
    while not res.Eof:
      for xcol in range(res.FieldCount):
        ws.write(xrow, xcol, res.GetFieldValueAt(xcol))
      xrow += 1
      res.Next()
    #save the xlsx
    wb.close()
    sw = returns.AddStreamWrapper()
    sw.LoadFromFile(tempfile)
  except:
    status.Err = str(sys.exc_info()[1])
    app.ConRead('a')
    
def SetData(config, params, returns):
  status = returns.CreateValues(['Err', ''])
  mlu = config.ModLibUtils
  app = config.AppObject
  app.ConCreate('out')
  tempfile = config.HomeDir + 'templates\\tmp_ls13.xlsx'
  config.BeginTransaction()
  try:
    if params.StreamWrapperCount > 0:
        sw = params.GetStreamWrapper(0)
    else:
        raise Exception, 'PERINGATAN!. Download stream not found'
    sw.SaveToFile(tempfile)
    wb = xloader(tempfile)
    app.ConWriteln('Cleanup existing rekening data...')
    s = '''
          delete from pbstmp.tmp_ls13
    '''
    config.ExecSQL(s)
    app.ConWriteln('Cleanup existing agunan data...')
    s = '''
          delete from pbstmp.tmp_ls13_agunan
    '''
    config.ExecSQL(s)
    app.ConWriteln('Reading rekening data...')
    ws = wb.get_sheet_by_name(name = 'Rekening')
    app.ConWriteln('getting column info (rekening)..')
    s = '''
          select column_name, data_type
          from all_tab_columns 
          where 
          table_name = 'TMP_LS13'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    collist = {}
    colstr = '('
    while not res.Eof:
      collist[res.column_name] = res.data_type
      colstr += res.column_name + ', '
      res.Next()
    colstr=colstr.rstrip(', ') + ')'
    #app.ConWriteln(str(collist))
    xrow = 1
    testcell = ws.cell(row = xrow, column=0)
    testvalue = testcell.value
    app.ConWriteln('getting row data (rekening)..')
    while testvalue not in (None,'',0,'None'):
      xcol = 0
      valstr = '('
      for colname in collist.keys():
        getcell = ws.cell(row = xrow, column = xcol)
        getvalue = getcell.value
        if collist[colname] != 'NUMBER':
          getvalue = mlu.QuotedStr(str(getvalue))
        valstr += str(getvalue) + ', '
        xcol += 1 
      valstr=valstr.rstrip(', ') + ')'
      s = '''
        insert into pbstmp.tmp_ls13 {0} values {1} 
      '''.format(colstr, valstr)
      config.ExecSQL(s)
      xrow += 1
      testcell = ws.cell(row = xrow, column=0)
      testvalue = testcell.value
    app.ConWriteln('{0} row(s) data updated (rekening).'.format(xrow-1))
    app.ConWriteln('Reading agunan data...')
    ws = wb.get_sheet_by_name(name = 'Agunan')
    app.ConWriteln('getting column info (agunan)..')
    s = '''
          select column_name, data_type
          from all_tab_columns 
          where 
          table_name = 'TMP_LS13_AGUNAN'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    collist = {}
    colstr = '('
    while not res.Eof:
      collist[res.column_name] = res.data_type
      colstr += res.column_name + ', '
      res.Next()
    colstr=colstr.rstrip(', ') + ')'
    xrow = 1
    testcell = ws.cell(row = xrow, column=0)
    testvalue = testcell.value
    app.ConWriteln('getting row data (agunan)..')
    while testvalue not in (None,'',0,'None'):
      xcol = 0
      valstr = '('
      for colname in collist.keys():
        getcell = ws.cell(row = xrow, column = xcol)
        getvalue = getcell.value
        if collist[colname] != 'NUMBER':
          getvalue = mlu.QuotedStr(str(getvalue))
        valstr += str(getvalue) + ', '
        xcol += 1 
      valstr=valstr.rstrip(', ') + ')'
      s = '''
        insert into pbstmp.tmp_ls13_agunan {0} values {1} 
      '''.format(colstr, valstr)
      config.ExecSQL(s)
      xrow += 1
      testcell = ws.cell(row = xrow, column=0)
      testvalue = testcell.value
    app.ConWriteln('{0} row(s) data updated (agunan).'.format(xrow-1))
    config.Commit()
  except:
    config.Rollback()
    status.Err = str(sys.exc_info()[1])
    app.ConRead('a')

def GetCSVData(config, params, returns):
  status = returns.CreateValues(['Err', ''])
  mlu = config.ModLibUtils
  app = config.AppObject
  app.ConCreate('out')
  rtmp = config.HomeDir + 'templates\\rekening_ls13.csv'
  rname = 'rekening_ls13.csv'
  atmp = config.HomeDir + 'templates\\agunan_ls13.csv'
  aname = 'agunan_ls13.csv'
  ztmp = config.HomeDir + 'templates\\tmp_ls13.zip'
  tmploc = config.HomeDir + 'templates' 
  try:
    app.ConWriteln('Generating file...')
    rkfile = open(rtmp, 'w')
    app.ConWriteln('get column name for rekening')
    s = '''
          select column_name
          from all_tab_columns 
          where 
          table_name = 'TMP_LS13'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    rowheader = ''
    while not res.Eof:
      rowheader += res.column_name
      rowheader += ';'
      res.Next()
    rkfile.write(rowheader.rstrip(';') + '\n')
    app.ConWriteln('get rekening data')
    s = '''
      select * from pbstmp.tmp_ls13
    '''
    res = config.CreateSQL(s).RawResult
    while not res.Eof:
      rowcontent = '' 
      for xcol in range(res.FieldCount):
        rowcontent += str(res.GetFieldValueAt(xcol))
        rowcontent += ';'
      rkfile.write(rowcontent.rstrip(';')+'\n')
      res.Next()
    rkfile.close()
    app.ConWriteln('get column name for agunan')
    s = '''
          select column_name
          from all_tab_columns 
          where 
          table_name = 'TMP_LS13_AGUNAN'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    agufile = open(atmp, 'w')
    rowheader = ''
    while not res.Eof:
      rowheader += res.column_name
      rowheader += ';'
      res.Next()
    agufile.write(rowheader.rstrip(';')+'\n')
    app.ConWriteln('get agunan data')
    s = '''
      select * from pbstmp.tmp_ls13_agunan
    '''
    res = config.CreateSQL(s).RawResult
    while not res.Eof:
      rowcontent = ''
      for xcol in range(res.FieldCount):
        rowcontent += str(res.GetFieldValueAt(xcol))
        rowcontent += ';'
      agufile.write(rowcontent.rstrip(';')+'\n')
      res.Next()
    agufile.close()
    zf = zipfile.ZipFile(ztmp, mode='w')
    zf.write(rtmp, rname)
    zf.write(atmp, aname)
    zf.close()
    sw = returns.AddStreamWrapper()
    sw.LoadFromFile(ztmp)
  except:
    status.Err = str(sys.exc_info()[1])
    app.ConWriteln(status.Err)
    app.ConRead('a')
    
def SetCSVData(config, params, returns):
  status = returns.CreateValues(['Err', ''])
  mlu = config.ModLibUtils
  app = config.AppObject
  app.ConCreate('out')
  rtmp = config.HomeDir + 'templates\\rekening_ls13.csv'
  rname = 'rekening_ls13.csv'
  atmp = config.HomeDir + 'templates\\agunan_ls13.csv'
  aname = 'agunan_ls13.csv'
  ztmp = config.HomeDir + 'templates\\tmp_ls13.zip'
  tmploc = config.HomeDir + 'templates' 
  config.BeginTransaction()
  try:
    if params.StreamWrapperCount > 0:
        sw = params.GetStreamWrapper(0)
    else:
        raise Exception, 'PERINGATAN!. Download stream not found'
    sw.SaveToFile(ztmp)
    zf = zipfile.ZipFile(ztmp)
    zf.extract(rname, tmploc)
    zf.extract(aname, tmploc)
    zf.close()
    app.ConWriteln('Cleanup existing rekening data...')
    s = '''
          delete from pbstmp.tmp_ls13
    '''
    config.ExecSQL(s)
    app.ConWriteln('Cleanup existing agunan data...')
    s = '''
          delete from pbstmp.tmp_ls13_agunan
    '''
    config.ExecSQL(s)
    app.ConWriteln('Reading rekening data...')
    rkfile = open(rtmp)
    app.ConWriteln('getting column info (rekening)..')
    s = '''
          select column_name, data_type
          from all_tab_columns 
          where 
          table_name = 'TMP_LS13'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    collist = {}
    colstr = '('
    while not res.Eof:
      collist[res.column_name] = res.data_type
      colstr += res.column_name + ', '
      res.Next()
    colstr=colstr.rstrip(', ') + ')'
    #app.ConWriteln(str(collist))
    rkfile.readline()
    testvalue = rkfile.readline()
    xrow = 0
    app.ConWriteln('getting row data (rekening)..')
    while testvalue not in (None,'',0,'None'):
      xcol = 0
      vallist = testvalue.rstrip('\n').split(';')
      valstr = '('
      for colname in collist.keys():
        getvalue = vallist[xcol]
        if collist[colname] != 'NUMBER':
          getvalue = mlu.QuotedStr(str(getvalue))
        valstr += str(getvalue) + ', '
        xcol += 1 
      valstr=valstr.rstrip(', ') + ')'
      s = '''
        insert into pbstmp.tmp_ls13 {0} values {1} 
      '''.format(colstr, valstr)
      config.ExecSQL(s)
      xrow += 1
      testvalue = rkfile.readline()
    app.ConWriteln('{0} row(s) data updated (rekening).'.format(xrow))
    rkfile.close()
    app.ConWriteln('Reading agunan data...')
    agufile = open(atmp)
    app.ConWriteln('getting column info (agunan)..')
    s = '''
          select column_name, data_type
          from all_tab_columns 
          where 
          table_name = 'TMP_LS13_AGUNAN'
          ORDER BY COLUMN_ID
    '''
    res = config.CreateSQL(s).RawResult
    collist = {}
    colstr = '('
    while not res.Eof:
      collist[res.column_name] = res.data_type
      colstr += res.column_name + ', '
      res.Next()
    colstr=colstr.rstrip(', ') + ')'
    xrow = 0
    agufile.readline()
    testvalue = agufile.readline()
    app.ConWriteln('getting row data (agunan)..')
    while testvalue not in (None,'',0,'None'):
      xcol = 0
      vallist = testvalue.rstrip('\n').split(';')
      valstr = '('
      for colname in collist.keys():
        getvalue = vallist[xcol]
        if collist[colname] != 'NUMBER':
          getvalue = mlu.QuotedStr(str(getvalue))
        valstr += str(getvalue) + ', '
        xcol += 1 
      valstr=valstr.rstrip(', ') + ')'
      s = '''
        insert into pbstmp.tmp_ls13_agunan {0} values {1} 
      '''.format(colstr, valstr)
      config.ExecSQL(s)
      xrow += 1
      testvalue = agufile.readline()
    app.ConWriteln('{0} row(s) data updated (agunan).'.format(xrow))
    agufile.close()
    config.Commit()
  except:
    config.Rollback()
    status.Err = str(sys.exc_info()[1])
    app.ConRead('a')

