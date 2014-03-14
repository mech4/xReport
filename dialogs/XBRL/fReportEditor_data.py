import com.ihsan.foundation.pobjecthelper as phelper
import com.ihsan.labs.excel as excel
import com.ihsan.util.xmlio as xutil
import sys, shutil, os
import pyFlexcel
from openpyxl import load_workbook as xloader
import traceback
import datetime
DEBUG_MODE = False
SIMPLE_DEBUG = False
inapp = None
DBGFRMIDS = ('',)

def OpenReport(config, parameter, returns):
  # config: ISysConfig object
  # parameter: TPClassUIDataPacket
  # returnpacket: TPClassUIDataPacket (undefined structure)
  rec = parameter.FirstRecord
  DTSFormId = rec.DTSFormId     #to get dtsmap -> dtsmapquery
  DTSFileName = rec.DTSFileName
  DTSId = rec.DTSId
  pCode = rec.pCode
  period_id = rec.period_id
  branch_id = rec.branch_id
  recflag = rec.recflag or 0
  mlu = config.ModLibUtils
  app = config.AppObject
  app.ConCreate('out')
  storeDir  = config.HomeDir+'data\\DTS\\'
  instanceDir = config.HomeDir+'data\\instance\\'
  status = returns.CreateValues(
      ['Is_Err', '']
  )
  helper = phelper.PObjectHelper(config)
  config.BeginTransaction()
  try:
    #cek periode
    s = '''
      select * from dtsreport where dtsid={0} and period_id={1} and branch_id={2}
    '''.format(
      str(DTSId),
      str(period_id),
      str(branch_id)
    )
    pCek = config.CreateSQL(s).RawResult
    # cek instance path
    iPath = instanceDir + str(DTSId)
    # if not exist create instance path
    if not os.path.exists(iPath):
      os.makedirs(iPath)
    # if report for period not exists
    s = '''select a.templatelocation||b.dtsfoldername tmp,
           a.dtslocation||b.dtsfoldername loc from dts a, dtsfolder b
           where a.dtsid=b.dtsid 
           and b.parentid is null 
           and a.dtsid={0}'''.format(str(DTSId))
    res = config.CreateSQL(s).RawResult
    tmpLocation = res.tmp
    dtsLocation = res.loc
    periodPath = iPath + '\\' + pCode + '\\' + str(branch_id)
    ipDTS = periodPath + '\\dts'
    ipData = periodPath + '\\data'
    #app.ConWriteln(dtsLocation)
    #app.ConWriteln(tmpLocation)
    
    if pCek.Eof:
      # create structure for current periode
      if not os.path.exists(periodPath):
        os.makedirs(periodPath)
        # create structure for instance(xbrl) and data(xls) in current periode
        os.makedirs(ipDTS)
        os.makedirs(ipData)
        shutil.copytree(dtsLocation, ipDTS + '\\' + dtsLocation.split('\\')[-1])
        shutil.copytree(tmpLocation, ipData + '\\' + tmpLocation.split('\\')[-1])
      # create report instance
      iReport = helper.CreatePObject('DTSReport')
      iReport.DTSId = DTSId
      iReport.period_id = period_id
      iReport.ReportLocation = periodPath
      iReport.branch_id = branch_id
    rf = xutil.XMLFolder()
    rf.setRoot(dtsLocation, False)
    cForm = rf.findFile(DTSFileName, True)
    if len(cForm) < 1:
      raise Exception, 'File {0} not found on DTS.'.format(DTSFileName)
    cForm = cForm[0]
    xsdPath = cForm.getFullPath()
    xlsPath = xsdPath.replace(dtsLocation, ipData + '\\' + tmpLocation.split('\\')[-1])
    sFileName = xlsPath.replace('.xsd','.xlsx')
    tmpSourceLoc = sFileName.replace(ipData + '\\' + tmpLocation.split('\\')[-1], tmpLocation)
    #app.ConWriteln('{0} from {1}'.format(sFileName,tmpSourceLoc))
    #app.ConRead(' ')
    #auto entry here
    # - check form edit flag (check using .done)
    flagFileName = sFileName.replace(sFileName.split('\\')[-1], '.done')
    if recflag == 1 and os.path.exists(flagFileName):
      os.remove(flagFileName)
      shutil.copy2(tmpSourceLoc, sFileName)
    if not os.path.exists(flagFileName): 
      # - check map type
      s = '''
        select * from dtsmap where dtsformid=%s
      ''' % str(DTSFormId)
      dtsmap = config.CreateSQL(s).RawResult
      ## if automatic
      if not dtsmap.Eof:
        s = '''
          select * from dtsmapquery where dtsmapid=%s
        ''' % dtsmap.dtsmapid
        # - get stored query string
        mapquery = config.CreateSQL(s).RawResult
        ## if location found
        if not mapquery.Eof:
          qryLoc = mapquery.querystring
          qryFile = open(qryLoc)
          qryText = qryFile.read()
          qryFile.close()
          #file ok            
          #preparing parameters
          s = '''
            select period_code from period where period_id=%s
          ''' % str(period_id)
          reportPeriod = config.CreateSQL(s).RawResult.period_code
          reportDate = (int(reportPeriod[-4:] or 0), int(reportPeriod[-6:-4] or 0), int(reportPeriod[-8:-6] or 0))
          startDate = mlu.EncodeDate(reportDate[0], reportDate[1], 1)
          sdstr = mlu.DecodeDate(startDate)
          startDateStr = '{0}-{1}-{2}'.format(str(sdstr[0]).zfill(4), str(sdstr[1]).zfill(2), str(sdstr[2]).zfill(2)) 
          if reportDate[1]<12:
            endDate = mlu.EncodeDate(reportDate[0], reportDate[1]+1, 1)-1
          else: 
            endDate = mlu.EncodeDate(reportDate[0]+1, 1, 1)-1
          edstr = mlu.DecodeDate(endDate)
          endDateStr = '{0}-{1}-{2}'.format(str(edstr[0]).zfill(4), str(edstr[1]).zfill(2), str(edstr[2]).zfill(2)) 
          s = '''
            select * from branchmember where branch_id=%s
          ''' % str(branch_id)
          branches = config.CreateSQL(s).RawResult
          branchList = ''
          while not branches.Eof:
            branchList += mlu.QuotedStr(branches.kode_cabang)
            branchList += ', '
            branches.Next()
          branchList = branchList.rstrip(', ')
          qryParam = {
            '_startdate' : mlu.QuotedStr(startDateStr),
            '_enddate' : mlu.QuotedStr(endDateStr),
            '_branchlist' : branchList, 
          }
          qryData = qryText
          for varkey in qryParam.keys():
            qryData = qryData.replace(varkey, qryParam[varkey])
          fetchedData = config.CreateSQL(qryData).RawResult
          #preparing datafile
          sFileWorkBook = xloader(sFileName)
          sFileWorkSheet = sFileWorkBook.get_active_sheet()
          #set row col param for any form type
          fType = config.CreateSQL('select formtype from dtsform where dtsformid=%s' % str(DTSFormId)).RawResult.formtype
          if fType == 'F':
            wrow = 1
            wcol = 2
          else:
            wrow = 2
            wcol = 0
          while not fetchedData.Eof:
            #write to xlsx here
            for fnum in range(fetchedData.FieldCount):
              if fType == 'F':
                currCell = sFileWorkSheet.cell(row=wrow+fnum, column=wcol)
              else:  
                currCell = sFileWorkSheet.cell(row=wrow, column=wcol+fnum)  
              currCell.value = fetchedData.GetFieldValueAt(fnum)
            if fType == 'F':
              wcol+=1
            else:
              wrow+=1
            fetchedData.Next()
          sFileWorkBook.save(sFileName)
          sFileWorkSheet = None
          sFileWorkBook = None
    #-- end of auto
    sw = returns.AddStreamWrapper()
    sw.LoadFromFile(sFileName)
    sw.MIMEType = 'xlsx'
    flagFile = open(flagFileName, 'w')
    flagFile.close()
    config.Commit()
  except:
    app.ConRead(' ')
    config.Rollback()
    status.Is_Err = str(sys.exc_info()[1])

  return 1

def SaveReport(config, parameter, returns):
  # config: ISysConfig object
  # parameter: TPClassUIDataPacket
  # returnpacket: TPClassUIDataPacket (undefined structure)
  rec = parameter.FirstRecord
  DTSFormId = rec.DTSFormId
  DTSFileName = rec.DTSFileName
  DTSId = rec.DTSId
  pCode = rec.pCode
  period_id = rec.period_id
  branch_id = rec.branch_id
  storeDir  = config.HomeDir+'data\\DTS\\'
  instanceDir = config.HomeDir+'data\\instance\\'
  status = returns.CreateValues(
      ['Is_Err', '']
  )
  helper = phelper.PObjectHelper(config)
  config.BeginTransaction()
  try:
    if parameter.StreamWrapperCount > 0:
        sw = parameter.GetStreamWrapper(0)
        sw.MIMEType = 'xlsx'
    else:
        raise Exception, 'PERINGATAN!. Download stream not found'
    iPath = instanceDir + str(DTSId)
    s = '''select a.templatelocation||b.dtsfoldername tmp,
           a.dtslocation||b.dtsfoldername loc from dts a, dtsfolder b
           where a.dtsid=b.dtsid 
           and b.parentid is null 
           and a.dtsid={0}'''.format(str(DTSId))
    res = config.CreateSQL(s).RawResult
    tmpLocation = res.tmp
    dtsLocation = res.loc
    periodPath = iPath + '\\' + pCode + '\\' + str(branch_id)
    ipDTS = periodPath + '\\dts'
    ipData = periodPath + '\\data'
    rf = xutil.XMLFolder()
    rf.setRoot(dtsLocation, False)
    cForm = rf.findFile(DTSFileName, True)
    if len(cForm) < 1:
      raise Exception, 'File {0} not found on DTS.'.format(DTSFileName)
    cForm = cForm[0]
    xsdPath = cForm.getFullPath()
    xlsPath = xsdPath.replace(dtsLocation, ipData + '\\' + tmpLocation.split('\\')[-1])
    sFileName = xlsPath.replace('.xsd','.xlsx')
    #if os.path.exists(sFileName):
    #  os.remove(sFileName)
    sw.SaveToFile(sFileName)
    config.Commit()
  except:
    config.Rollback()
    status.Is_Err = str(sys.exc_info()[1])

  return 1

def GenReport(config, parameter, returns):
    # config: ISysConfig object
    # parameter: TPClassUIDataPacket
    # returnpacket: TPClassUIDataPacket (undefined structure)
  def advSeek(rootSearch, tag, attrib, value):
    rbTag = rootSearch.seek(tag)
    rbAttr = []
    rbVal = []
    if attrib == '__text':
      for ele in rbTag:
        if ele.text == value:
          rbVal.append(ele)
    else:
      for ele in rbTag:
        if ele.attrib.has_key(attrib):
          rbAttr.append(ele)
      for ele in rbAttr:
        if ele.attrib[attrib] == value:
          rbVal.append(ele)
    return rbVal
  #--

  def getTag(rootSearch, attrib, value, app=None):
    if app!=None:
      app.ConWriteln('search tag on element with %s = %s' % (attrib,value))
    unfilter = rootSearch.seek(value, 'attrib', True)
    rbAttr = []
    for ele in unfilter:
      if app!=None:
        app.ConWriteln('found %s' % ele.tag)
      if ele.attrib.has_key(attrib):
        rbAttr.append(ele)
        if app!=None:
          app.ConWriteln('has attrib')
      else:
        if app!=None:
          app.ConWriteln('no attrib')
    for ele in rbAttr:
      if ele.attrib[attrib] == value and ele.qTag != 'link':
        if app!=None:
          app.ConWriteln('with qTag %s' % ele.qTag)
        return ele.tag
    return None
  #--
  def checkIVFormula(instance, app):
    ## Currently support flat form only (single context)
    ## update 'T' & 'M' supported
    linkbases = instance.schema.linkbases['formula']
    if len(linkbases) < 1:
      #raise Exception, 'Formula linkbases not found on this schema.'
      app.ConWriteln('Formula linkbases not found on this schema.')
    else:
      linkbase = linkbases[0]
    linkbase.readFromFile()
    IVFormulas = linkbase.rootElement.seek('generalVariable')
    LinkPool = []
    for formulaElement in IVFormulas:
      if formulaElement.attrib.has_key('select'):
        checkstr = formulaElement.attrib['select'] 
        if '$in' in checkstr:
           foundLink = checkstr.split('$in')[1].split(' ')[0]
           if foundLink not in LinkPool:
             LinkPool.append(foundLink)
    return LinkPool
  #--
  def eaCheck(root, rumus, mType, app=None):
    varlist = rumus[2]
    for key in varlist.keys():
      cvar = key + ' '
      csou = varlist[key]
      result = root.seek(csou)
      if app:
        app.ConWriteln('Checking %s for existance' % csou)
      if len(result)==0:
        if app:
          app.ConWriteln('Status : Not Exists')
        return 0
      else:
        if app:
          app.ConWriteln('Status : Exists')
        pass
    return 1
  #--
  def calcResult(root, rumus, mType, app=None):
    frm = rumus[0] + ' '
    varlist = rumus[2]
    for key in varlist.keys():
      cvar = key + ' '
      csou = varlist[key]
      valueContainers = root.seek(csou)
      for vc in valueContainers:
        frm=frm.replace('$%s' % cvar, vc.text)
    if frm.find('$') > -1:
      if app:
        app.ConWriteln('Unable to calculate : %s' % frm)
      return 0
    else:
      res = eval(frm)
      newEle = xutil.XMLElement(rumus[3], 'base')
      newEle.text = str(int(res))
      newEle.attrib['contextRef'] = 'c1'
      newEle.attrib['unitRef'] = 'I'
      newEle.attrib['decimals'] = 2
      pos = len(root.childrens)
      if mType=='F':
        pos -= 2
      root.append(newEle, pos)
      if app:
        app.ConWriteln('Added %s' % rumus[3])
      return 1 
  #--
  def vaResult(root, rumus, mType, app=None):
    frm = rumus[0] + ' '
    frm=frm.replace('number', 'int')
    varlist = rumus[2]
    if frm.count('=')==2 and frm.count('==')==1:
      frm = frm.replace('==','=')
    leftside, rightside = frm.split('=')
    for key in varlist.keys():
      cvar = key + ' '
      csou = varlist[key]
      if cvar in rightside:
        valueContainers = root.seek(csou)
        for vc in valueContainers:
          if csou[0] in ('s','d') and csou!='dummy':
            rightside=rightside.replace('$%s' % cvar, '"'+str(vc.text)+'"')
          else:
            rightside=rightside.replace('$%s' % cvar, str(vc.text))
      else:
        if csou[0] in ('s','d') and csou!='dummy':
          valueContainers = root.seek(csou)
          for vc in valueContainers:
            leftside=leftside.replace('$%s' % cvar, '"'+str(vc.text)+'"')
        else:
          leftside=leftside.replace('$%s' % cvar, csou)
    if rightside.find('$') > -1 or leftside.find('$') > -1:
      if app:
        app.ConWriteln('Variable not fully resolved :')
        app.ConWriteln('Formula : %s' % rightside)
        app.ConWriteln('Varlist : %s' % str(varlist))
      return 0
    else:
      if len(leftside.replace('(','').replace(')','').strip())>6:
        if app:
          app.ConWriteln('Assignment : {0} <- {1}'.format(leftside,rightside))
          app.ConWriteln('Not really assignment, skipped for test.')
        return 1
      else:
        res = eval(rightside)
        newEle = xutil.XMLElement(leftside.strip(), 'base')
        newEle.text = str(int(res))
        newEle.attrib['contextRef'] = 'c1'
        newEle.attrib['unitRef'] = 'I'
        newEle.attrib['decimals'] = 2
        pos = len(root.childrens)
        if mType=='F':
          pos -= 2
        root.append(newEle, pos)
        if app:
          app.ConWriteln('Entry added : {0} = {1}'.format(leftside.strip(),str(int(res))))
        return 1
  #-- 
  def vaCheck(root, rumus, mType, app=None):
    frm = str(rumus[0]) + ' '
    #manfix rumus
    #frm=frm.replace(' = ',' == ')
    fele = frm.split('=')
    if len(fele)>1:
      ftidy = ''
      can_use = True
      for i in fele:
        if len(i)<1:
          can_use = False
      if can_use:
        for i in range(len(fele)):
          thisPart = fele[i].strip() 
          if i>0:
            prevPart = fele[i-1].strip()
            if prevPart[-1] not in ('>','<','!','=') and thisPart[0] not in ('>','<','!','='):
              ftidy+=' == '+thisPart
            else:
              ftidy+='= '+thisPart
          else:
            ftidy+=thisPart
        if app:
          app.ConWriteln('Tidy Formula : %s' % ftidy)
        frm = ftidy
    #end of manfix rumus '=' menjadi '=='
    frm=frm.replace('number', 'int')
    frm=frm.replace('> ==','>=')
    varlist = rumus[2]
    linklist = {}
    sumlist = []
    for key in varlist.keys():
      csou = varlist[key]
      VarIsLink = True if "/" in csou or "period-instant" in csou else False
      linklist[key] = VarIsLink
    #begin sum process
    #remove sum for iv leave for local
    if 'sum' in frm:
      nfrm = ''
      sumparts = frm.split('sum')
      for eachpart in sumparts:
        if sumparts.index(eachpart)>0:
          sumvar = eachpart.split('(',1)[-1].split(')',1)[0].strip().strip('$')
          oldstate = '('+eachpart.split('(',1)[-1].split(')',1)[0]+')'
          if linklist[sumvar]:
            newstate = ' '+eachpart.split('(',1)[-1].split(')',1)[0].strip()+' '
            nfrm += eachpart.replace(oldstate,newstate)
          else:
            if '[' in varlist[sumvar]:
              vSource = varlist[sumvar]
              tableName = vSource.split('[')[0].strip() 
              cPhrase = vSource[vSource.find('[')+1:vSource.find(']')]
              vSource = vSource.split(':')[-1]
              #if app:
              #  app.ConWriteln('get sum of {0} \r\n{2} value with condition {1}'.format(varlist[sumvar],cPhrase,vSource))
              nfrm += eachpart.replace(oldstate,str(int(get_sum(root, vSource, tableName, cPhrase, app))))
            else:
              nfrm += eachpart.replace(oldstate,str(int(get_sum(root, varlist[sumvar], app=app))))
        else:
          nfrm += eachpart
      frm = nfrm
      if app:
        app.ConWriteln('Sum fix : %s' % frm)
    #end sum process
    #add extra space to formula
    frm = frm + ' ' 
    for key in varlist.keys():
      skey = str(key)
      cvar = skey + ' '
      b = frm.split(skey)
      for c in range(len(b)):
        if len(b[c])>0:
          if not b[c][0].isdigit():
            if b[c][0] != ' ':
              b[c] = ' '+b[c]
      frm = skey.join(b)
      csou = varlist[key]
      VarIsLink = linklist[key] 
      if VarIsLink:
        LinkCode = csou.find('(')
        if LinkCode == 3:
          LinkVal = get_doc(csou, app)
          frm=frm.replace('$%s' % cvar, str(LinkVal))
        elif LinkCode == 5:
          LinkVal = get_count(root)
          frm=frm.replace('$%s' % cvar, str(LinkVal))
          if app:
            app.ConWriteln('Count Test (bypass):')
            app.ConWriteln(frm)
            #app.ConRead('Checking count.')
          frm = '1==1' #count not used
        elif LinkCode == 7:
          LinkVal = '"{0}-{1}-{2}"'.format(thn,bln,tgl)
          frm=frm.replace('$%s' % cvar, LinkVal)
        elif LinkCode == 9:
          LinkVal = get_substring(app)
          frm=frm.replace('$%s' % cvar, str(LinkVal))
        else:
          csou = csou.split(':')[-1]
          valueContainers = root.seek(csou)
          for vc in valueContainers:
              if csou[0] in ('s','d') and csou!='dummy':
                frm=frm.replace('$%s' % cvar, '"'+str(vc.text)+'"')
              else:
                frm=frm.replace('$%s' % cvar, str(vc.text))
      else:
        if '[' in csou:
          csou = csou.split(':')[-1]
          # non sum condition ..?
          '''
          condition = csou[csou.find('[')+1:csou.find(']')]
          tname = csou.split('[')[0].strip()
          vsource = csou.split(':')[-1]
          cFields = []
          #reformat condition phrase
          condition = condition.replace('eq','=')
          condition = condition.replace('=','==')
          condition = condition.replace('(',' ( ')
          condition = condition.replace(')',' ) ')
          tidyCondition = condition
          if 'len' not in condition:
            step1 = condition.split('==')
            for idx in range(len(step1)):
              if idx>0:
                step2 = step1[idx].lstrip()
                tidyCondition = tidyCondition.replace(' '+step2.split(' ')[0]+' ',' "{0}" '.format(step2.split(' ')[0]))
          #get lookups condition fields
          tidyCondition = tidyCondition.replace('ne','!=')
          tidyCondition = tidyCondition.replace('ge','>=')
          tidyCondition = tidyCondition.replace('string-length','strln')
          tidyCondition = tidyCondition.replace('len','strln')
          tidyCondition = tidyCondition.replace('le','<=')
          tidyCondition = tidyCondition.replace('number','int')
          tidyCondition = tidyCondition.replace('strln','len')
          tidyCondition = tidyCondition.replace('""','"')
          lFields = condition.split('base:')
          for idx in range(len(lFields)):
            if idx>0 and lFields[idx].split(' ')[0] not in cFields:
              cFields.append(lFields[idx].split(' ')[0])
          #create valueContainer = {row1 : [sum_field, cond1, cond2, .., condN], row2 : [sum_field, cond1, cond2, .., condN], .., rowN : [sum_field, cond1, cond2, .., condN]}
          #ga jd create valueContainer, diganti realtime sum
          if app:
            app.ConWriteln('Tidy Condition : %s' % tidyCondition)
          testCondition = tidyCondition
          for look in cFields:
            res = root.seek(look, category='tag', exact=True, fromRoot=False)
            if len(res)>0:
              res = res[0]
              if app:
                #app.ConWriteln('Find %s value in : \r\n%s' % (look,res.writeXML(False)))
                app.ConWriteln('Found %s as value of %s.' % (res.text,look))
              if look[0] == 's':
                testCondition = testCondition.replace('base:'+look+' ', '"{0}" '.format(res.text))
              else:
                testCondition = testCondition.replace('base:'+look+' ', '{0} '.format(res.text))
          bfTest = True
          testCondition = testCondition+' '
          testPassed = None
          while bfTest:
            try:
              testPassed = eval("True if "+testCondition+" else False")
              bfTest = False
            except:
              if 'defined' in str(sys.exc_info()[1]):
                culprit = str(sys.exc_info()[1]).split("name '")[-1].split("'")[0]
                testCondition = testCondition.replace(' {0} '.format(culprit),' "{0}" '.format(culprit))
                bfTest = True
              else:
                bfTest = False
          if app:
            app.ConWriteln('Test Condition : %s' % testCondition)
            app.ConWriteln('Result : %s' % str(testPassed))
          if testPassed:
            getvalue = root.seek(sum_field, category='tag', exact=True, fromRoot=False)
            if len(getvalue)>0:
              result_element = getvalue[0]
              if app:
                #app.ConWriteln('Get %s value in : \r\n%s' % (sum_field,result_element.writeXML(False)))
                app.ConWriteln('Found %s as value of %s.' % (result_element.text,sum_field))
              sum_value += int(result_element.text)
          '''
          #--end non sum condition 
        valueContainers = root.seek(csou)
        for vc in valueContainers:
            if csou[0] in ('s','d') and csou!='dummy':
              frm=frm.replace('$%s' % cvar, '"'+str(vc.text)+'"')
            else:
              frm=frm.replace('$%s' % cvar, str(vc.text))
    if frm.find('$') > -1:
      if app:
        app.ConWriteln('Variable not fully resolved : ')
        app.ConWriteln('Formula : {0}\r\nVarlist : {1}'.format(frm, str(varlist)))
      if SIMPLE_DEBUG and inapp:
        inapp.ConWriteln('Variable not fully resolved : ')
        inapp.ConWriteln('Formula : {0}\r\nVarlist : {1}'.format(frm, str(varlist)))
      return 0
    else:
      try:
        res = eval(frm)
      except:
        raise Exception, "Error formula :\r\n%s\r\nFrom : \r\n%s" % (frm, str(rumus[0]) + ' ')
      if res:
        if app:
          app.ConWriteln('Test passed : %s' % frm)
        return 2
      if app:
        app.ConWriteln('Test not passed : %s' % frm)
      if SIMPLE_DEBUG and inapp:
        inapp.ConWriteln('Test not passed : %s' % frm)
      return 1
  #-- 
  #IV section dev
  def get_count(rootSearch):
    eContainer = rootSearch.seek('contextRef', category='attrib', exact=True, fromRoot=True)
    cCount = len(eContainer)
    return str(cCount)
  #--
  def get_substring(app=None):
    if app:
      app.ConWriteln('Get Indetifier : %s' % bCode[0:3])
    return '"{0}"'.format(bCode[0:3]) 
  #--  
  def get_date():
    return thn,bln,tgl
  #--
  def get_doc(vlink, app=None):
    #fix untidy vlink
    vlink = vlink.split('&')[0]
    if vlink.find('[')>0:
      use_condition = True
    else:
      use_condition = False
    vtarget = vlink.split('$in')[1].split(' ')[0]
    targetCol = vlink.split(':')[-1]
    targetRoot = IVInstance[vtarget].rootElement
    if use_condition:
      #parseCondition(vlink)
      cPhrase = vlink[vlink.find('[')+1:vlink.find(']')]
      if app:
        app.ConWriteln('doc with condition %s from %s' % (cPhrase, vlink))
      result = str(int(get_sum(targetRoot, targetCol, vtarget, cPhrase, app)))
    else:
      if vtarget in vlink.split('xbrli:xbrl')[-1]:
        #csum
        result = str(int(get_sum(targetRoot, targetCol, vtarget, app=app)))
      else:
        #single
        res = targetRoot.seek(targetCol)
        if len(res)>0:
         res = res[0]
        else:
         raise Exception, 'Field {0} on {1} not found.'.format(targetCol, vtarget)
        result = res.text
    return result
  #--

  def get_sum(rootSearch, sum_field, table_name=None, condition=None, app=None):
    #define initial sum_value
    sum_value = 0
    if condition:
      rows = rootSearch.seek(table_name, category='tag', exact=True, fromRoot=True)
      if app:
        #app.ConWriteln('Find %s value in : \r\n%s' % (table_name, rootSearch.writeXML()))
        app.ConWriteln('Found %s rows.' % str(len(rows)))
      cFields = []
      #reformat condition phrase
      condition = condition.replace('eq','=')
      condition = condition.replace('=','==')
      condition = condition.replace('(',' ( ')
      condition = condition.replace(')',' ) ')
      tidyCondition = condition
      if 'len' not in condition:
        step1 = condition.split('==')
        for idx in range(len(step1)):
          if idx>0:
            step2 = step1[idx].lstrip()
            tidyCondition = tidyCondition.replace(' '+step2.split(' ')[0]+' ',' "{0}" '.format(step2.split(' ')[0]))
      #get lookups condition fields
      tidyCondition = tidyCondition.replace('ne','!=')
      tidyCondition = tidyCondition.replace('ge','>=')
      tidyCondition = tidyCondition.replace('string-length','strln')
      tidyCondition = tidyCondition.replace('len','strln')
      tidyCondition = tidyCondition.replace('le','<=')
      tidyCondition = tidyCondition.replace('number','int')
      tidyCondition = tidyCondition.replace('strln','len')
      tidyCondition = tidyCondition.replace('""','"')
      lFields = condition.split('base:')
      for idx in range(len(lFields)):
        if idx>0 and lFields[idx].split(' ')[0] not in cFields:
          cFields.append(lFields[idx].split(' ')[0])
      #create valueContainer = {row1 : [sum_field, cond1, cond2, .., condN], row2 : [sum_field, cond1, cond2, .., condN], .., rowN : [sum_field, cond1, cond2, .., condN]}
      #ga jd create valueContainer, diganti realtime sum
      if app:
        app.ConWriteln('Tidy Condition : %s' % tidyCondition)
        app.ConWriteln('cFields : %s' % str(cFields))
      for row in rows:
        testCondition = tidyCondition
        for look in cFields:
          res = row.seek(look, category='tag', exact=True, fromRoot=False)
          if len(res)>0:
            res = res[0]
            if app:
              #app.ConWriteln('Find %s value in : \r\n%s' % (look,res.writeXML(False)))
              app.ConWriteln('Found %s as value of %s.' % (res.text,look))
            if look[0] == 's':
              testCondition = testCondition.replace('base:'+look+' ', '"{0}" '.format(res.text))
            else:
              testCondition = testCondition.replace('base:'+look+' ', '{0} '.format(res.text))
        bfTest = True
        testCondition = testCondition+' '
        testPassed = None
        while bfTest:
          try:
            testPassed = eval("True if "+testCondition+" else False")
            bfTest = False
          except:
            if 'defined' in str(sys.exc_info()[1]):
              culprit = str(sys.exc_info()[1]).split("name '")[-1].split("'")[0]
              testCondition = testCondition.replace(' {0} '.format(culprit),' "{0}" '.format(culprit))
              bfTest = True
            else:
              bfTest = False
        if app:
          app.ConWriteln('Test Condition : %s' % testCondition)
          app.ConWriteln('Result : %s' % str(testPassed))
          app.ConRead(' ')
        if testPassed:
          getvalue = row.seek(sum_field, category='tag', exact=True, fromRoot=False)
          if len(getvalue)>0:
            result_element = getvalue[0]
            if app:
              #app.ConWriteln('Get %s value in : \r\n%s' % (sum_field,result_element.writeXML(False)))
              app.ConWriteln('Found %s as value of %s.' % (result_element.text,sum_field))
            sum_value += int(result_element.text) 
    else:
      rows = rootSearch.seek(sum_field, category='tag', exact=True, fromRoot=True)
      if app:
        #app.ConWriteln('Find %s value in : \r\n%s' % (sum_field, rootSearch.writeXML()))
        app.ConWriteln('Found %s rows.' % str(len(rows)))
      for row in rows:
        sum_value += int(row.text)
    #-- endif
    return sum_value
  #--
  
  def get_none(vlink):
    return vlink.split(':')[-1]
  #--
  def xlint2date(xldate, datemode=0):
    # datemode: 0 for 1900-based, 1 for 1904-based
    return (
        datetime.datetime(1899, 12, 30)
        + datetime.timedelta(days=xldate + 1462 * datemode)
        )
  #--
  rec = parameter.FirstRecord
  DTSFormId = rec.DTSFormId
  DTSFileName = rec.DTSFileName
  DTSFormCode = rec.DTSFormCode
  DTSId = rec.DTSId
  pCode = rec.pCode
  bCode = rec.bCode
  formType = rec.FormType
  IsEmpty = rec.IsEmpty
  dataSize = rec.dataSize
  app = config.AppObject
  app.ConCreate('out')
  global inapp
  if SIMPLE_DEBUG:
    inapp = app
  period_id = rec.period_id
  branch_id = rec.branch_id
  storeDir  = config.HomeDir+'data\\DTS\\'
  instanceDir = config.HomeDir+'data\\instance\\'
  status = returns.CreateValues(
      ['Is_Err', ''],
      ['fName', ''],
      ['is_valid','']
  )
  calcValidation = 0
  global DEBUG_MODE
  helper = phelper.PObjectHelper(config)
  mlu = config.ModLibUtils
  config.BeginTransaction()
  try:
    if IsEmpty != 'T':
      if parameter.StreamWrapperCount > 0:
          sw = parameter.GetStreamWrapper(0)
      else:
          raise Exception, 'PERINGATAN!. Download stream not found'
    app.ConWriteln('Reading Taxonomy')
    iPath = instanceDir + str(DTSId)
    s = '''select a.templatelocation||b.dtsfoldername tmp,
           a.dtslocation||b.dtsfoldername loc from dts a, dtsfolder b
           where a.dtsid=b.dtsid 
           and b.parentid is null 
           and a.dtsid={0}'''.format(str(DTSId))
    res = config.CreateSQL(s).RawResult
    tmpLocation = res.tmp
    dtsLocation = res.loc
    periodPath = iPath + '\\' + pCode + '\\' + str(branch_id)
    bln = int(pCode[:2])
    thn = int(pCode[2:])
    if bln<12:
      nextmo = mlu.EncodeDate(thn, bln+1, 1)
    else:
      nextmo = mlu.EncodeDate(thn+1, 1, 1)
    reportdate = nextmo-1
    thn, bln, tgl = mlu.DecodeDate(reportdate)
    str_repdate = '{0}-{1}-{2}'.format(str(thn).zfill(4),str(bln).zfill(2),str(tgl).zfill(2))
    ipDTS = periodPath + '\\dts'
    ipData = periodPath + '\\data'
    rf = xutil.XMLFolder()
    rf.setRoot(ipDTS, False)
    ### update versioning here
    s = '''
      select * from dtsdict where dtsid=%s
    ''' % str(DTSId)
    dictlist = config.CreateSQL(s).RawResult
    while not dictlist.Eof:
      rf.addDicts(dictlist.dictname, dictlist.dictloc)
      dictlist.Next()
    ### --
    cForm = rf.findFile(DTSFileName, True)
    if len(cForm) < 1:
      raise Exception, 'File {0} not found on DTS.'.format(DTSFileName)
    cForm = cForm[0]
    xsdPath = cForm.getFullPath()
    app.ConWriteln('Reading form schema')
    sForm = xutil.xbrlSchema(cForm.fileName, cForm.folder)
    s = '''
      select * from dtsmeta where dtsformid=%s
    ''' % str(DTSFormId)
    formMeta = config.CreateSQL(s).RawResult
    metaPool = {}
    orphan = []
    nillchecker = {}
    enumchecker = {}
    if not formMeta.Eof:
      #construct meta into pool
      while not formMeta.Eof:
        newMeta = xutil.metaTree(formMeta.metaname, formMeta.metaqname)
        metaPool[formMeta.dtsmetaid] = (newMeta, formMeta.metaparent)
        hasValue = True
        if formMeta.metatype == 'Empty':
          hasValue = False
        nillable = True
        if formMeta.nillable in ('F','f'):
          nillable = False
        nillchecker[formMeta.metaname] = nillable
        if formMeta.metaenum not in (None,'','None',0):
          enumchecker[formMeta.metaname] = formMeta.metaenum
        newMeta.define(
          datatype=formMeta.metatype,
          description=formMeta.metadesc,
          enum=formMeta.metaenum or None,
          nillable=nillable,
          hasValue=hasValue 
        )
        formMeta.Next()
      #rearange parent in pool
      for metaid in metaPool.keys():
        parentid = metaPool[metaid][1]
        xmeta = metaPool[metaid][0]  
        if parentid not in (None,'','None'):
          metaParent = metaPool[parentid][0]
          metaParent.append(xmeta)
        else:
          orphan.append(xmeta)
      #define meta root
      if len(orphan)>1:
        rootname = DTSFileName.split('.',1)[0]
        metaRoot = xutil.metaTree(rootname,rootname)
        for newChild in orphan:
          metaRoot.append(newChild)
      else:
        metaRoot = orphan[0]
      sForm.metaStructure = metaRoot
    #Prepare enum used
    enumList = {}
    s = '''
      select * from dtsenum where dtsid={0} and dtsenumname in (
        select distinct(metaenum) from dtsmeta where dtsformid={1} and metaenum is not null
      )
    '''.format(str(DTSId), str(DTSFormId))
    enumUsed = config.CreateSQL(s).RawResult
    while not enumUsed.Eof:
      enumName = enumUsed.dtsenumname
      enumCode = enumUsed.dtsenumvalue
      enumDesc = enumUsed.dtsenumdesc
      if enumName not in enumList.keys():
        enumList[enumName] = {}
      enumList[enumName][enumCode] = enumDesc
      enumUsed.Next()
    #--
    sForm.mappingType = formType
    xlsPath = xsdPath.replace(dtsLocation, ipData + '\\' + tmpLocation.split('\\')[-1])
    sFileName = xlsPath.replace('.xsd','.xlsx')
    app.ConWriteln('Initializing form instance')
    if bCode in ('',None,0):
      bCode = '517001000'
    if len(bCode)==6:
      bCode = bCode+'000'
    if len(bCode)!=9 or not bCode.isdigit():
      raise Exception, 'Kode wilayah %s tidak sesuai format standar' % str(bCode)
    if DEBUG_MODE:
      app.ConWriteln('DEBUG MODE')
      app.ConWriteln('----------')
    if SIMPLE_DEBUG and not DEBUG_MODE:
      app.ConWriteln('DEBUG MODE [SIMPLE]')
      app.ConWriteln('----------')
    if DEBUG_MODE or SIMPLE_DEBUG:
      app.ConWriteln('Report Code : %s' % DTSFormCode)
      app.ConWriteln('Report Date : %s' % str_repdate)
      app.ConWriteln('Kode Cabang : %s' % bCode)
    iForm = xutil.xbrlInstance(sForm, bCode, str_repdate)
    if DEBUG_MODE:
      app.ConWriteln('- Meta Tree')
      if IsEmpty != 'T':
        app.ConWriteln(str(iForm.readMeta()))
      else:
        app.ConWriteln('Empty form ignore meta tree')
    # temporary constant define here
    cxId = 'c1'
    contextParam = { cxId :('PBS',bCode,str_repdate)}
    ##
    ### update root here ###
    iRoot = iForm.rootElement
    iRoot.namespace['xmlns:xsi'] = "http://www.w3.org/2001/XMLSchema-instance"
    ### update v22, interformvalidation link
    IVPool = checkIVFormula(iForm, app)
    IVPool.append('parameters')
    IVInstance = {}
    if len(IVPool)>0:
      s = '''
        select a.dtsaliaslink, replace(b.dtsfilename, '.xsd') lCode, b.dtsfilename from dtsalias a, dtsfile b
        where a.dtsaliasloc = b.dtsfileid and replace(b.dtsfilename, '.xsd') in (%s) and dtsid=%s 
      ''' % (str(IVPool).replace('[','').replace(']',''), str(DTSId))
      aliases = config.CreateSQL(s).RawResult
      while not aliases.Eof:
        Loc = aliases.lCode
        LinkLoc = aliases.dtsaliaslink
        LinkLoc = LinkLoc.replace('%20','_')
        LinkLoc = LinkLoc.replace('/'+Loc+'.xsd', '') 
        iRoot.namespace['xmlns:%s' % 'par' if Loc=='parameters' else Loc] = LinkLoc
        if Loc !='parameters':
          if DEBUG_MODE:
            app.ConWriteln('Loc Check')
            app.ConWriteln('---------')
          ivFileName = str(bCode)+ '-' + str_repdate + '-MM-' + Loc + '.xml'
          ivFile = rf.findFile(ivFileName, True)
          if len(ivFile)>0:
            ivFile = ivFile[0]
            ivLoc = ivFile.getFullPath()
            ivFile.readFromFile()
            if DEBUG_MODE:
              app.ConWriteln(ivLoc)
              #app.ConWriteln(ivFile.rootElement.writeXML())
            IVInstance[Loc] = ivFile
          else:
            raise Exception, 'Validasi antar form gagal, instance %s belum tersedia.' % Loc
        aliases.Next()
    else:
      app.ConWriteln('Interform Validation formula link not needed')
    ###--
    aliasIdx = rf.aliases.values().index(sForm)
    sLoc = rf.aliases.keys()[aliasIdx]
    iRoot.namespace['xsi:schemaLocation'] = "{0} {1}".format(
      sLoc.replace('%20','_').replace('/'+DTSFileName,''),
      sLoc)  
    iForm.addContext(contextParam)
    iForm.addFooter()
    iForm.addDummy()
    accerr = ''
    if IsEmpty != 'T':
      if os.path.exists(sFileName):
        os.remove(sFileName)
      sw.SaveToFile(sFileName)
      app.ConWriteln('opening template...')
      owb = xloader(sFileName, data_only=True)
      app.ConWriteln('get active sheet...')
      ows = owb.get_active_sheet()
      app.ConWriteln('Reading instance data')
      #running through validation
      aercount = 1
      if formType == 'F':
        xlskey = ows.cell(row=1, column=0).value
        xlsElement = []
        xrow = 1
        xlsStartCol = 2
        while xlskey not in (None,'','None'):
          xlsElement.append((xrow, ows.cell(row=xrow, column=0).value))
          xrow +=1
          xlskey = ows.cell(row=xrow, column=0).value
        contentParam = {}
        contentParam[cxId] = {'ID_1' : {}} 
        for elementRow in xlsElement:
          thisValue = ows.cell(row=elementRow[0], column=xlsStartCol).value
          if thisValue in (None,'',"None"):
            #check nillable and enum skipped for flat
            thisValue = '0'
          if str(thisValue).replace('.','').replace('e+','').isdigit():
            if thisValue == int(thisValue):
              thisValue = int(thisValue) 
          contentParam[cxId]['ID_1'][elementRow[1]] = str(thisValue) or '0'
        app.ConWriteln('Preparing instance in memory')
        iForm.addContent(contentParam)
        #######ADD FORMULA HERE !!!!!!!!!##############
        # run pre-va formula
        vfSkipFlag = 0
        # read va formula
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,  
          
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='v' and b.varsource is not null 
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId) # and b.vartype<>'l'
        res = config.CreateSQL(s).RawResult
        vflist = {}
        while not res.Eof:
          #vflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor, res.vartype]
          vflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = vflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        # exec pre-va formula
        for vf in vflist.keys():
          #trap assignment       
          if 'if' not in vflist[vf][0] and '<' not in vflist[vf][0] and '>' not in vflist[vf][0] and 'sum' not in vflist[vf][0] and '/' not in str(vflist[vf][2]) and '(' not in str(vflist[vf][2]):
            if len(vflist[vf][0].split('=')) == 2 or (vflist[vf][0].count('=')==2 and vflist[vf][0].count('==')==1):
              if vflist[vf][0].split('=')[0].count('$') == 1:
                if DEBUG_MODE:
                  app.ConWriteln('FID : %s' % str(vf))
                vares = vaResult(iForm.rootElement, vflist[vf], formType, app if DEBUG_MODE or str(vf) in DBGFRMIDS else None)
                if vares<1:
                  if SIMPLE_DEBUG:
                    app.ConWriteln('FID : %s [1]' % str(vf))
                  if DEBUG_MODE:
                    app.ConWriteln("Skipped assignment #{2} : {0} for {1}".format(vflist[vf][0], vflist[vf][3], str(vf)))
                  vfSkipFlag += 1
        # run calc formula
        cfSkipFlag = 0
        # read calc formula
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,  
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='c' and b.varsource is not null
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId)
        res = config.CreateSQL(s).RawResult
        cflist = {}
        while not res.Eof:
          cflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = cflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        # exec calc formula
        for cf in cflist.keys():
          if DEBUG_MODE:
            app.ConWriteln('FID : %s' % str(cf))
          calcres = calcResult(iForm.rootElement, cflist[cf], formType, app if DEBUG_MODE or str(cf) in DBGFRMIDS else None)
          if calcres<1:
            if SIMPLE_DEBUG:
              app.ConWriteln('FID : %s [1]' % str(cf))
            #if DEBUG_MODE:
            #  app.ConWriteln("skipped calculation #{2} : {0} for {1}".format(cflist[cf][0], cflist[cf][3], str(cf)))
            cfSkipFlag += 1
        # run post-va formula
        vfSkipFlag = 0
        # exec post-va formula
        for vf in vflist.keys():
          #run all
          if DEBUG_MODE:
            app.ConWriteln('FID : %s' % str(vf))
          vares = vaCheck(iForm.rootElement, vflist[vf], formType, app if DEBUG_MODE or str(vf) in DBGFRMIDS else None)
          if vares<1:
            if SIMPLE_DEBUG:
              app.ConWriteln('FID : %s [1]' % str(vf))
            #raise Exception, 'Unable to calculate'
            #skip IV
            #app.ConWriteln('IV skipped (dev)')
            break
          if vares<2: 
            if SIMPLE_DEBUG:
              app.ConWriteln('FID : %s [1]' % str(vf))
            #raise Exception, vflist[vf][1] # raise error message
            accerr += '{0}. '.format(str(aercount).rjust(5))
            errmsg = vflist[vf][1]
            accerr += errmsg or ''
            accerr += '\r\n' 
            aercount += 1
        # run ea formula
        # read ea formula
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,  
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='e' and b.varsource is not null
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId)
        res = config.CreateSQL(s).RawResult
        eflist = {}
        while not res.Eof:
          eflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = eflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        # exec ea formula
        for ef in eflist.keys():
          if DEBUG_MODE:
            app.ConWriteln('FID : %s' % str(ef))
          eares = eaCheck(iForm.rootElement, eflist[ef], formType, app if DEBUG_MODE or str(ef) in DBGFRMIDS else None)
          if eares<1:
            if SIMPLE_DEBUG:
              app.ConWriteln('FID : %s [1]' % str(ef))
            #raise Exception, eflist[ef][1] # raise error message
            accerr += '{0}. '.format(str(aercount).rjust(5))
            accerr += vflist[vf][1] or ''
            accerr += '\r\n' 
            aercount += 1
      elif formType == 'T':
        xlskey = ows.cell(row=0, column=0).value
        xlsElement = []
        xcol = 0
        xlsStartRow = 2
        contentNum = 1
        xrow = xlsStartRow 
        valueTester = ows.cell(row=xrow, column=0).value
        if SIMPLE_DEBUG:
          inapp = None
        #######READ FORMULA FOR ROW HERE !!!!!!!!!##############
        # read va formula (r)
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,  
          
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='v' and b.varsource is not null and a.exectype='r'
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId)
        res = config.CreateSQL(s).RawResult
        vflist = {}
        while not res.Eof:
          vflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = vflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        # read calc formula (r)
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='c' and b.varsource is not null and a.exectype='r'
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId)
        res = config.CreateSQL(s).RawResult
        cflist = {}
        while not res.Eof:
          cflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = cflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        # read ea formula (r)
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,  
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='e' and b.varsource is not null and a.exectype='r'
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId)
        res = config.CreateSQL(s).RawResult
        eflist = {}
        while not res.Eof:
          eflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = eflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        while xlskey not in (None,'','None'):
          xlsElement.append((xcol, ows.cell(row=0, column=xcol).value))    
          xcol +=1
          xlskey = ows.cell(row=0, column=xcol).value
        while valueTester not in (None,'','None'):
          contentId = 'ID_{0}'.format(str(contentNum))
          contentParam = {}
          contentRow = { contentId : {}}
          for elementCol in xlsElement:
            thisValue = ows.cell(row=xrow, column=elementCol[0]).value
            #enum validation
            if elementCol[1] in enumchecker.keys() and not nillchecker[elementCol[1]]:
              enumName = enumchecker[elementCol[1]]  
              if str(thisValue) not in enumList[enumName].keys():
                erm = '(Record ID : {0}) Field {1} harus berisi kode {2}.'.format(str(contentNum),elementCol[1],enumName) 
                #raise Exception, erm # raise error message
                accerr += '{0}. '.format(str(aercount).rjust(5))
                accerr += erm
                accerr += '\r\n' 
                aercount += 1
            if thisValue in (None,'',"None"):
              #nillable validation
              if not nillchecker[elementCol[1]]:
                erm = '(Record ID : {0}) Field {1} tidak boleh kosong.'.format(str(contentNum),elementCol[1]) 
                #raise Exception, erm # raise error message
                accerr += '{0}. '.format(str(aercount).rjust(5))
                accerr += erm
                accerr += '\r\n' 
                aercount += 1
              if elementCol[1][0] == 'm':
                thisValue = '0'
              else:
                thisValue = ''
            if elementCol[1][0] == 'd': #if is date
              if type(thisValue) != type('string'): 
                if str(thisValue).isdigit():
                  thisValue = xlint2date(thisValue)
                #if date is not read asstring
                #create string from encoded date
                thisValue = thisValue.isoformat()[:10]        
              else:
                thisValue = thisValue[:10] #trim string if string is timestamp not date
            if str(thisValue).replace('.','').isdigit():
              if thisValue == int(thisValue):
                thisValue = int(thisValue) 
            contentRow[contentId][elementCol[1]] = str(thisValue)
          contentParam[cxId] = contentRow 
          iForm.addContent(contentParam)
          testRoot = advSeek(iForm.rootElement, DTSFormCode, 'id', contentId)[0]
          # exec pre-va formula (r)
          for vf in vflist.keys():
            #trap assignment
            if 'if' not in vflist[vf][0] and '<' not in vflist[vf][0] and '>' not in vflist[vf][0] and 'sum' not in vflist[vf][0] and '/' not in str(vflist[vf][2]) and '(' not in str(vflist[vf][2]):
              if len(vflist[vf][0].split('=')) == 2 or (vflist[vf][0].count('=')==2 and vflist[vf][0].count('==')==1):
                if vflist[vf][0].split('=')[0].count('$') == 1:
                  if DEBUG_MODE:
                    app.ConWriteln('FID : %s' % str(vf))
                  vares = vaResult(testRoot, vflist[vf], formType, app if DEBUG_MODE or str(vf) in DBGFRMIDS else None)
                  if vares<1:
                    if SIMPLE_DEBUG:
                      app.ConWriteln('FID : %s [1]' % str(vf))
                    #if DEBUG_MODE:
                    #  app.ConWriteln("skipped assignment #{2} : {0} for {1}".format(vflist[vf][0], vflist[vf][3], str(vf)))
                    vfSkipFlag += 1
          # exec calc formula (r)
          for cf in cflist.keys():
            if DEBUG_MODE:
              app.ConWriteln('FID : %s' % str(cf))
            calcres = calcResult(testRoot, cflist[cf], formType, app if DEBUG_MODE or str(cf) in DBGFRMIDS else None)
            if calcres<1:
              if SIMPLE_DEBUG:
                app.ConWriteln('FID : %s [1]' % str(cf))
              #if DEBUG_MODE:
              #  app.ConWriteln("skipped calculation #{2} : {0} for {1}".format(cflist[cf][0], cflist[cf][3], str(cf)))
              cfSkipFlag += 1
          # exec post-va formula (r)
          for vf in vflist.keys():
            #run all
            if DEBUG_MODE:
              app.ConWriteln('FID : %s' % str(vf))
            vares = vaCheck(testRoot, vflist[vf], formType, app if DEBUG_MODE or str(vf) in DBGFRMIDS else None)
            if vares<1:
              #if SIMPLE_DEBUG:
              #  app.ConWriteln('FID : %s [1]' % str(vf))
              #if DEBUG_MODE:
              #  app.ConWriteln(str(vflist[vf]))
              #app.ConWriteln('IV skipped (dev)')
              break
            if vares<2: 
              if SIMPLE_DEBUG:
                app.ConWriteln('FID : %s [2]' % str(vf))
              erm = vflist[vf][1]
              erm = erm.replace('{$v1/../@id}', str(contentNum)) 
              #raise Exception, erm # raise error message
              accerr += '{0}. '.format(str(aercount).rjust(5))
              accerr += erm
              accerr += '\r\n' 
              aercount += 1
          # exec ea formula (r)
          for ef in eflist.keys():
            if DEBUG_MODE:
              app.ConWriteln('FID : %s' % str(ef))
            eares = eaCheck(testRoot, eflist[ef], formType, app if DEBUG_MODE or str(ef) in DBGFRMIDS else None)
            if eares<1:
              if SIMPLE_DEBUG:
                app.ConWriteln('FID : %s [1]' % str(ef))
              #raise Exception, eflist[ef][1] # raise error message
              accerr += '{0}. '.format(str(aercount).rjust(5))
              accerr += eflist[ef][1] or ''
              accerr += '\r\n' 
              aercount += 1
          xrow+=1
          valueTester = ows.cell(row=xrow, column=0).value
          contentNum+=1
        #######ADD FORMULA HERE !!!!!!!!!(Porting from flat to table)##############
        # run pre-va formula
        vfSkipFlag = 0
        # read va formula
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,
            
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='v' and b.varsource is not null and a.exectype<>'r'
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId)
        res = config.CreateSQL(s).RawResult
        vflist = {}
        while not res.Eof:
          vflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = vflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        # exec pre-va formula
        if SIMPLE_DEBUG:
          inapp = app
        for vf in vflist.keys():
          #trap assignment
          #app.ConWriteln(str(vflist[vf]))
          if 'if' not in vflist[vf][0] and '<' not in vflist[vf][0] and '>' not in vflist[vf][0] and 'sum' not in vflist[vf][0] and '/' not in str(vflist[vf][2]) and '(' not in str(vflist[vf][2]):
            if len(vflist[vf][0].split('=')) == 2 or (vflist[vf][0].count('=')==2 and vflist[vf][0].count('==')==1):
              if vflist[vf][0].split('=')[0].count('$') == 1:
                if DEBUG_MODE:
                  app.ConWriteln('FID : %s' % str(vf))
                vares = vaResult(iForm.rootElement, vflist[vf], formType, app if DEBUG_MODE or str(vf) in DBGFRMIDS else None)
                if vares<1:
                  if SIMPLE_DEBUG:
                    app.ConWriteln('FID : %s [1]' % str(vf))
                  #if DEBUG_MODE:
                  #  app.ConWriteln("Skipped assignment #{2} : {0} for {1}".format(vflist[vf][0], vflist[vf][3], str(vf)))
                  vfSkipFlag += 1
        # run calc formula
        cfSkipFlag = 0
        # read calc formula
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,  
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='c' and b.varsource is not null and a.exectype='f'
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId)
        res = config.CreateSQL(s).RawResult
        cflist = {}
        while not res.Eof:
          cflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = cflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        # exec calc formula
        for cf in cflist.keys():
          if DEBUG_MODE:
            app.ConWriteln('FID : %s' % str(cf))
          calcres = calcResult(iForm.rootElement, cflist[cf], formType, app if DEBUG_MODE or str(cf) in DBGFRMIDS else None)
          if calcres<1:
            if SIMPLE_DEBUG:
              app.ConWriteln('FID : %s [1]' % str(cf))
            #if DEBUG_MODE:
            #  app.ConWriteln("skipped calculation #{2} : {0} for {1}".format(cflist[cf][0], cflist[cf][3], str(cf)))
            cfSkipFlag += 1
        # run post-va formula
        vfSkipFlag = 0
        # exec post-va formula
        for vf in vflist.keys():
          #run all
          #if DEBUG_MODE:
          #  app.ConWriteln(str(vflist))
          if DEBUG_MODE:
            app.ConWriteln('FID : %s' % str(vf))
          vares = vaCheck(iForm.rootElement, vflist[vf], formType, app if DEBUG_MODE or str(vf) in DBGFRMIDS else None) 
          if vares<1:
            if SIMPLE_DEBUG:
              app.ConWriteln('FID : %s [1]' % str(vf))
            #app.ConWriteln('IV skipped (dev)')
            break
          if vares<2: 
            if SIMPLE_DEBUG:
              app.ConWriteln('FID : %s [2]' % str(vf))
            #raise Exception, vflist[vf][1] # raise error message
            accerr += '{0}. '.format(str(aercount).rjust(5))
            accerr += vflist[vf][1] or ''
            accerr += '\r\n' 
            aercount += 1
        # run ea formula
        # read ea formula
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,  
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='e' and b.varsource is not null and a.exectype='f'
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId)
        res = config.CreateSQL(s).RawResult
        eflist = {}
        while not res.Eof:
          eflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = eflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        # exec ea formula
        for ef in eflist.keys():
          if DEBUG_MODE:
            app.ConWriteln('FID : %s' % str(ef))
          eares = eaCheck(iForm.rootElement, eflist[ef], formType, app if DEBUG_MODE or str(ef) in DBGFRMIDS else None)
          if eares<1:
            if SIMPLE_DEBUG:
              app.ConWriteln('FID : %s [1]' % str(ef))
            #raise Exception, eflist[ef][1] # raise error message
            accerr += '{0}. '.format(str(aercount).rjust(5))
            accerr += eflist[ef][1] or ''
            accerr += '\r\n' 
            aercount += 1
      elif formType == 'M':
        schemaMeta = iForm.readMeta()
        tstmeta = iForm.schema.metaStructure
        #app.ConWriteln(str(tstmeta.getMeta(False)))
        #app.ConWriteln(str(schemaMeta))
        subName = None
        for lv, entity, QN, dType, mDesc in schemaMeta:
          if lv==1 and (dType=='Table' or dType=='Empty'):
            subName = entity
        if subName == None:
          raise Exception, 'Cannot determine sub entry name.'
        def getMetaLv(schemaMeta, param):
          rlv = None
          for lv, entity, QN, dType, mDesc in schemaMeta:
            if entity == param:
              rlv = lv
          return rlv
        xlskey = ows.cell(row=0, column=0).value
        xlsElement = []
        xcol = 0
        xlsStartRow = 2
        contentNum = 1
        xrow = xlsStartRow
        valueTester = ows.cell(row=xrow, column=0).value
        #######READ FORMULA FOR ROW HERE !!!!!!!!!##############
        # read va formula (r)
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,  
          
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='v' and b.varsource is not null and a.exectype='r'
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId)
        res = config.CreateSQL(s).RawResult
        vflist = {}
        while not res.Eof:
          vflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = vflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        # read calc formula (r)
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,  
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='c' and b.varsource is not null and a.exectype='r'
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId)
        res = config.CreateSQL(s).RawResult
        cflist = {}
        while not res.Eof:
          cflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = cflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        # read ea formula (r)
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,  
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='e' and b.varsource is not null and a.exectype='r'
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId)
        res = config.CreateSQL(s).RawResult
        eflist = {}
        while not res.Eof:
          eflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = eflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        while xlskey not in (None,'','None'):
          xlsElement.append((xcol, ows.cell(row=0, column=xcol).value, getMetaLv(schemaMeta, ows.cell(row=0, column=xcol).value)))    
          xcol +=1
          xlskey = ows.cell(row=0, column=xcol).value
        LastRow = None
        while valueTester not in (None,'','None'):
          if DEBUG_MODE:
            app.ConWriteln(str(valueTester))
          if valueTester != LastRow:
            subCount = 1
            contentId = 'ID_{0}'.format(str(contentNum))
            contentNum+=1
            contentParam = {}
            contentRow = { contentId : {}}
            for elementCol in xlsElement:
              #app.ConWriteln("{0} ; {1}".format(elementCol[1],elementCol[2]))
              thisValue = ows.cell(row=xrow, column=elementCol[0]).value
              #enum validation
              if elementCol[1] in enumchecker.keys() and not nillchecker[elementCol[1]]:
                enumName = enumchecker[elementCol[1]]  
                if str(thisValue) not in enumList[enumName].keys():
                  erm = '(Record ID : {0}) Field {1} harus berisi kode {2}.'.format(str(contentNum-1),elementCol[1],enumName) 
                  #raise Exception, erm # raise error message
                  accerr += '{0}. '.format(str(aercount).rjust(5))
                  accerr += erm
                  accerr += '\r\n' 
                  aercount += 1
              if thisValue in (None,'',"None"):
                #nillable validation
                if not nillchecker[elementCol[1]]:
                  erm = '(Record ID : {0}) Field {1} tidak boleh kosong.'.format(str(contentNum-1),elementCol[1]) 
                  #raise Exception, erm # raise error message
                  accerr += '{0}. '.format(str(aercount).rjust(5))
                  accerr += erm
                  accerr += '\r\n' 
                  aercount += 1
                if elementCol[1][0] == 'm':
                  thisValue = '0'
                else:
                  thisValue = ''
              if elementCol[1][0] == 'd': #if is date
                if type(thisValue) != type('string'): 
                  if str(thisValue).isdigit():
                    thisValue = xlint2date(thisValue)
                  #if date is not read asstring
                  #create string from encoded date
                  thisValue = thisValue.isoformat()[:10]        
                else:
                  thisValue = thisValue[:10] #trim string if string is timestamp not date
              if str(thisValue).replace('.','').isdigit():
                if thisValue == int(thisValue):
                  thisValue = int(thisValue) 
              if elementCol[2]==2:
                if not contentRow[contentId].has_key(subName):
                  if DEBUG_MODE:
                    app.ConWriteln('creating subname %s' % subName)
                  contentRow[contentId][subName] = {}
                else:
                  if contentRow[contentId][subName].__class__.__name__ != 'dict':
                    contentRow[contentId][subName] = {}
                if not contentRow[contentId][subName].has_key(subCount):
                  if DEBUG_MODE:
                    app.ConWriteln('creating subcount %s' % str(subCount))
                  contentRow[contentId][subName][subCount] = {}
                contentRow[contentId][subName][subCount][elementCol[1]] = str(thisValue)
              else:
                contentRow[contentId][elementCol[1]] = str(thisValue)
          else:
            for elementCol in xlsElement:
              thisValue = ows.cell(row=xrow, column=elementCol[0]).value
              #enum validation
              if elementCol[1] in enumchecker.keys() and not nillchecker[elementCol[1]]:
                enumName = enumchecker[elementCol[1]]  
                if str(thisValue) not in enumList[enumName].keys():
                  erm = '(Record ID : {0}) Field {1} harus berisi kode {2}.'.format(str(contentNum-1),elementCol[1],enumName) 
                  #raise Exception, erm # raise error message
                  accerr += '{0}. '.format(str(aercount).rjust(5))
                  accerr += erm
                  accerr += '\r\n' 
                  aercount += 1
              if thisValue in (None,'',"None"):
                #nillable validation
                if not nillchecker[elementCol[1]]:
                  erm = '(Record ID : {0}) Field {1} tidak boleh kosong.'.format(str(contentNum-1),elementCol[1]) 
                  #raise Exception, erm # raise error message
                  accerr += '{0}. '.format(str(aercount).rjust(5))
                  accerr += erm
                  accerr += '\r\n' 
                  aercount += 1
                if elementCol[1][0] == 'm':
                  thisValue = '0'
                else:
                  thisValue = ''
              if elementCol[1][0] == 'd': #if is date
                if type(thisValue) != type('string'): 
                  if str(thisValue).isdigit():
                    thisValue = xlint2date(thisValue)
                  #if date is not read asstring
                  #create string from encoded date
                  thisValue = thisValue.isoformat()[:10]        
                else:
                  thisValue = thisValue[:10] #trim string if string is timestamp not date
              if str(thisValue).replace('.','').isdigit():
                if thisValue == int(thisValue):
                  thisValue = int(thisValue) 
              if elementCol[2]==2:
                if not contentRow[contentId].has_key(subName):
                  contentRow[contentId][subName] = {}
                if not contentRow[contentId][subName].has_key(subCount):
                  contentRow[contentId][subName][subCount] = {}
                contentRow[contentId][subName][subCount][elementCol[1]] = str(thisValue)
              else:
                contentRow[contentId][elementCol[1]] = str(thisValue)
          LastRow = valueTester
          xrow+=1
          valueTester = ows.cell(row=xrow, column=0).value
          subCount+=1
          if valueTester != LastRow:
            contentParam[cxId] = contentRow
            iForm.addContent(contentParam)
            #app.ConWriteln(str(len(advSeek(iForm.rootElement, DTSFormCode, 'id', contentId))))
            #app.ConWriteln(iForm.rootElement.writeXML())
            testRoot = advSeek(iForm.rootElement, DTSFormCode, 'id', contentId)[0]
            # exec pre-va formula (r)
            for vf in vflist.keys():
              #trap assignment
              if 'if' not in vflist[vf][0] and '<' not in vflist[vf][0] and '>' not in vflist[vf][0] and 'sum' not in vflist[vf][0] and '/' not in str(vflist[vf][2]) and '(' not in str(vflist[vf][2]):
                if len(vflist[vf][0].split('=')) == 2 or (vflist[vf][0].count('=')==2 and vflist[vf][0].count('==')==1):
                  if vflist[vf][0].split('=')[0].count('$') == 1:
                    if DEBUG_MODE:
                      app.ConWriteln('FID : %s' % str(vf))
                    vares = vaResult(testRoot, vflist[vf], formType, app if DEBUG_MODE or str(vf) in DBGFRMIDS else None)
                    if vares<1:
                      if SIMPLE_DEBUG:
                        app.ConWriteln('FID : %s [1]' % str(vf))
                      #if DEBUG_MODE:
                      #  app.ConWriteln("Skipped assignment #{2} : {0} for {1}".format(vflist[vf][0], vflist[vf][3], str(vf)))
                      vfSkipFlag += 1
            # exec calc formula (r)
            for cf in cflist.keys():
              if DEBUG_MODE:
                app.ConWriteln('FID : %s' % str(cf))
              calcres = calcResult(testRoot, cflist[cf], formType, app if DEBUG_MODE or str(cf) in DBGFRMIDS else None)
              if calcres<1:
                if SIMPLE_DEBUG:
                  app.ConWriteln('FID : %s [1]' % str(cf))
                #if DEBUG_MODE:
                #  app.ConWriteln("Skipped calculation #{2} : {0} for {1}".format(cflist[cf][0], cflist[cf][3], str(cf)))
                cfSkipFlag += 1
            # exec post-va formula (r)
            for vf in vflist.keys():
              #run all
              if DEBUG_MODE:
                app.ConWriteln('FID : %s' % str(vf))
              vares = vaCheck(testRoot, vflist[vf], formType, app if DEBUG_MODE or str(vf) in DBGFRMIDS else None)
              if vares<1:
                if SIMPLE_DEBUG:
                  app.ConWriteln('FID : %s [1]' % str(vf))
                #app.ConWriteln('IV skipped (dev)')
                break
              if vares<2:
                if SIMPLE_DEBUG:
                  app.ConWriteln('FID : %s [2]' % str(vf))
                errstr = vflist[vf][1].replace('{$v1/../@id}', str(contentNum-1)) 
                #raise Exception, errstr # raise error message
                accerr += '{0}. '.format(str(aercount).rjust(5))
                accerr += errstr
                accerr += '\r\n' 
                aercount += 1
            # exec ea formula (r)
            for ef in eflist.keys():
              if DEBUG_MODE:
                app.ConWriteln('FID : %s' % str(ef))
              eares = eaCheck(testRoot, eflist[ef], formType, app if DEBUG_MODE or str(ef) in DBGFRMIDS else None)
              if eares<1:
                if SIMPLE_DEBUG:
                  app.ConWriteln('FID : %s [1]' % str(ef))
                #raise Exception, eflist[ef][1] # raise error message
                accerr += '{0}. '.format(str(aercount).rjust(5))
                accerr += eflist[ef][1] or ''
                accerr += '\r\n' 
                aercount += 1
        #######ADD FORMULA HERE !!!!!!!!!(Porting from table to multi)##############
        # run pre-va formula
        vfSkipFlag = 0
        # read va formula
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='v' and b.varsource is not null and a.exectype<>'r'
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId)
        res = config.CreateSQL(s).RawResult
        vflist = {}
        while not res.Eof:
          vflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = vflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        # exec pre-va formula
        for vf in vflist.keys():
          #trap assignment
          if 'if' not in vflist[vf][0] and '<' not in vflist[vf][0] and '>' not in vflist[vf][0] and 'sum' not in vflist[vf][0] and '/' not in str(vflist[vf][2]) and '(' not in str(vflist[vf][2]):
            if len(vflist[vf][0].split('=')) == 2 or (vflist[vf][0].count('=')==2 and vflist[vf][0].count('==')==1):
              if vflist[vf][0].split('=')[0].count('$') == 1:
                if DEBUG_MODE:
                  app.ConWriteln('FID : %s' % str(vf))
                vares = vaResult(iForm.rootElement, vflist[vf], formType, app if DEBUG_MODE or str(vf) in DBGFRMIDS else None)
                if vares<1:
                  if SIMPLE_DEBUG:
                    app.ConWriteln('FID : %s [1]' % str(vf))
                  #if DEBUG_MODE:
                  #  app.ConWriteln("Skipped assignment #{2} : {0} for {1}".format(vflist[vf][0], vflist[vf][3], str(vf)))
                  vfSkipFlag += 1
        # run calc formula
        cfSkipFlag = 0
        # read calc formula
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,  
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='c' and b.varsource is not null and a.exectype='f'
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId)
        res = config.CreateSQL(s).RawResult
        cflist = {}
        while not res.Eof:
          cflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = cflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        # exec calc formula
        for cf in cflist.keys():
          if DEBUG_MODE:
            app.ConWriteln('FID : %s' % str(cf))
          calcres = calcResult(iForm.rootElement, cflist[cf], formType, app if DEBUG_MODE or str(cf) in DBGFRMIDS else None)
          if calcres<1:
            if SIMPLE_DEBUG:
              app.ConWriteln('FID : %s [1]' % str(cf))
            #if DEBUG_MODE:
            #  app.ConWriteln("skipped calculation #{2} : {0} for {1}".format(cflist[cf][0], cflist[cf][3], str(cf)))
            cfSkipFlag += 1
        # run post-va formula
        vfSkipFlag = 0
        # exec post-va formula
        for vf in vflist.keys():
          #run all
          if DEBUG_MODE:
            app.ConWriteln('FID : %s' % str(vf))
          vares = vaCheck(iForm.rootElement, vflist[vf], formType, app if DEBUG_MODE or str(vf) in DBGFRMIDS else None)
          #vares = 3
          if vares<1:
            if SIMPLE_DEBUG:
              app.ConWriteln('FID : %s [1]' % str(vf))
            #app.ConWriteln('IV skipped (dev)')
            break
          if vares<2: 
            if SIMPLE_DEBUG:
              app.ConWriteln('FID : %s [2]' % str(vf))
            #raise Exception, vflist[vf][1] # raise error message
            accerr += '{0}. '.format(str(aercount).rjust(5))
            accerr += vflist[vf][1] or ''
            accerr += '\r\n' 
            aercount += 1
        # run ea formula
        # read ea formula
        s = '''
          select 
          a.dtsformulaid, 
          a.rumus, 
          a.applyfor,
          a.message,  
          listagg(b.varname||'='||b.varsource, ',,') within group (order by b.varid) as xvars
          from dtsformula a, dtsformulavars b
          where a.dtsformulaid=b.dtsformulaid and a.dtsformid=%s and a.formulatype='e' and b.varsource is not null and a.exectype='f'
          group by a.dtsformulaid, a.rumus, a.applyfor, a.message
        ''' % str(DTSFormId)
        res = config.CreateSQL(s).RawResult
        eflist = {}
        while not res.Eof:
          eflist[res.dtsformulaid] = [res.rumus, res.message, {}, res.applyfor]
          varlist = eflist[res.dtsformulaid][2]
          for v in res.xvars.split(',,'):
            varlist[v.split('=',1)[0]] = v.split('=',1)[1]
          res.Next()
        # exec ea formula
        for ef in eflist.keys():
          if DEBUG_MODE:
            app.ConWriteln('FID : %s' % str(ef))
          eares = eaCheck(iForm.rootElement, eflist[ef], formType, app if DEBUG_MODE or str(ef) in DBGFRMIDS else None)
          if eares<1:
            if SIMPLE_DEBUG:
              app.ConWriteln('FID : %s [1]' % str(ef))
            #raise Exception, eflist[ef][1] # raise error message
            accerr += '{0}. '.format(str(aercount).rjust(5))
            accerr += eflist[ef][1]
            accerr += '\r\n' 
            aercount += 1
      else:
          pass
    ixPath = xsdPath.replace(dtsLocation, ipDTS + '\\' + tmpLocation.split('\\')[-1])
    iFileName = str(bCode)+ '-' + str_repdate + '-MM-' + DTSFileName.replace('.xsd','.xml')
    ixPath = ixPath.replace(DTSFileName, iFileName)
    if accerr != '':
      app.ConWriteln('Tidak lolos validasi bisnis.')
      status.is_valid = 'err'
      raise Exception, accerr
    else:
      app.ConWriteln('Preparing output file')  
      if DEBUG_MODE:
        app.ConWriteln('@ %s' % ixPath)
      iFile = open(ixPath, "w")
      iFile.write(iForm.rootElement.writeXML())
      iFile.close()
      sw = returns.AddStreamWrapper()
      sw.LoadFromFile(ixPath)
      status.fName = iFileName
    ## test run
    #raise Exception, 'Adding xbrl instance content function unfinished.'
    if calcValidation > 0:
      app.ConWriteln('calcValidation : %s' % str(calcValidation))
      app.ConRead('calcValidation : %s' % str(calcValidation))
    if DEBUG_MODE or SIMPLE_DEBUG:
      app.ConWriteln('No Error. End of Debug. Cancel to Rollback. Enter to Proceed.')
      app.ConRead('No Error. End of test.')
    config.Commit()
  except:
    app.ConWriteln('Error : %s' % str(sys.exc_info()[1]))
    if DEBUG_MODE or SIMPLE_DEBUG or '' not in DBGFRMIDS:
      app.ConWriteln('Traceback Log')
      app.ConWriteln('-------------')
      _errmsg = traceback.format_exc().splitlines()
      for line in _errmsg : 
        app.ConWriteln(str(line))
      #app.ConWriteln(iForm.rootElement.writeXML())
      app.ConRead('Error')
    config.Rollback()
    status.Is_Err = str(sys.exc_info()[1]) + ' '

  return 1

def PrepareTemp(config, parameter, returns):
  # config: ISysConfig object
  # parameter: TPClassUIDataPacket
  # returnpacket: TPClassUIDataPacket (undefined structure)
  rec = parameter.FirstRecord
  DTSFormId = rec.DTSFormId
  DTSFileName = rec.DTSFileName
  DTSId = rec.DTSId
  pCode = rec.pCode
  period_id = rec.period_id
  branch_id = rec.branch_id
  storeDir  = config.HomeDir+'data\\DTS\\'
  instanceDir = config.HomeDir+'data\\instance\\'
  status = returns.CreateValues(
      ['Is_Err', '']
  )          
  helper = phelper.PObjectHelper(config)
  config.BeginTransaction()
  try:
    #cek periode
    s = '''
      select * from dtsreport where dtsid={0} and period_id={1} and branch_id={2}
    '''.format(
      str(DTSId),
      str(period_id),
      str(branch_id)
    )
    pCek = config.CreateSQL(s).RawResult
    # cek instance path
    iPath = instanceDir + str(DTSId)
    # if not exist create instance path
    if not os.path.exists(iPath):
      os.makedirs(iPath)
    # if report for period not exists
    s = '''select a.templatelocation||b.dtsfoldername tmp,
           a.dtslocation||b.dtsfoldername loc from dts a, dtsfolder b
           where a.dtsid=b.dtsid 
           and b.parentid is null 
           and a.dtsid={0}'''.format(str(DTSId))
    res = config.CreateSQL(s).RawResult
    tmpLocation = res.tmp
    dtsLocation = res.loc
    periodPath = iPath + '\\' + pCode + '\\' + str(branch_id)
    ipDTS = periodPath + '\\dts'
    ipData = periodPath + '\\data'
    rf = xutil.XMLFolder()
    rf.setRoot(dtsLocation, False)
    cForm = rf.findFile(DTSFileName, True)
    if len(cForm) < 1:
      raise Exception, 'File {0} not found on DTS.'.format(DTSFileName)
    cForm = cForm[0]
    xsdPath = cForm.getFullPath()
    xlsPath = xsdPath.replace(dtsLocation, ipData + '\\' + tmpLocation.split('\\')[-1])
    sFileName = xlsPath.replace('.xsd','.xlsx')
    if not os.path.exists(sFileName):
      tFileName = xsdPath.replace(dtsLocation, tmpLocation)
      tFileName = tFileName.replace('.xsd','.xlsx')
      shutil.copy(tFileName, sFileName)

    config.Commit()
  except:
    config.Rollback()
    status.Is_Err = str(sys.exc_info()[1])

  return 1

def PeriodHandler(config, params, returns):
  config.BeginTransaction()
  s = "select distinct periode_type from reportclass"
  res = config.CreateSQL(s).RawResult
  while not res.Eof:
    PeriodCheck(config, res.periode_type)
    res.Next()
  #--
  config.Commit()
  return
   
def PeriodCheck(config, ptype):
  def periodGenerate(period_type, tgl, bln, thn, hari):
    mon = ('', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December')
    qtr = ('', '1st Quarter', '2nd Quarter', '3rd Quarter', '4th Quarter')
    week = ('', '1st Week of', '2nd Week of', '3rd Week of', '4th Week of', '5th Week of')
    dayname = ('', 'Sunday,', 'Monday,', 'Tuesday,', 'Wednesday,', 'Thrusday,', 'Friday,', 'Saturday,')
    if period_type=='Y':
      return str(thn), str(thn)
    elif period_type=='M':
      return str(bln).zfill(2)+str(thn), mon[bln]+' '+str(thn)      
    elif period_type=='Q':
      return str((bln/4)+1).zfill(2)+str(thn), qtr[(bln/4)+1]+' '+str(thn)
    elif period_type=='W':
      if tgl<8:
        return str(thn)+str(bln).zfill(2)+'1', week[1]+' '+mon[bln]+' '+str(thn)
      elif tgl<16:
        return str(thn)+str(bln).zfill(2)+'2', week[2]+' '+mon[bln]+' '+str(thn)
      elif tgl<24:
        return str(thn)+str(bln).zfill(2)+'3', week[3]+' '+mon[bln]+' '+str(thn)
      else:
        return str(thn)+str(bln).zfill(2)+'4', week[4]+' '+mon[bln]+' '+str(thn)
    else:
      return str(tgl).zfill(2)+str(bln).zfill(2)+str(thn), dayname[hari]+' '+str(tgl).zfill(2)+' '+mon[bln]+' '+str(thn)     
  #--
  mlu = config.ModLibUtils
  app = config.AppObject
  app.ConCreate('out')
  tgl = mlu.DecodeDate(config.Now())
  hari = mlu.DayOfWeek(config.Now())
  bln = tgl[1]
  thn = tgl[0]
  tglnum = tgl[2]
  #app.ConWriteln('{0} : {1} : {2}'.format(ptype,str(tgl),str((bln/3)+1).zfill(2)))
  #app.ConRead('a')
  period = periodGenerate(ptype, tglnum, bln, thn, hari)
  #raise Exception, period[0]
  #app.ConWriteln(str(period[0]))
  s = "select * from period where period_code='%s' and period_type='%s'" % (period[0], ptype)
  res = config.CreateSQL(s).RawResult
  if not res.Eof:
    #raise Exception, 'Ada'
    pass
  else:
    #raise Exception, 'Belum Ada'
    s = "insert into period (period_id, period_code, description, period_type) values (seq_period.nextval, '%s', '%s', '%s')" % (period[0], period[1], ptype)
    #raise Exception, s
    config.ExecSQL(s)
  
  if ptype=='D' and hari!=6:
    selisih_hari = 6-hari
    hari=6
    tglnum = tglnum+selisih_hari
    if bln<12:
      nmbln = bln+1
      nmthn = thn
    else:
      nmbln = 1
      nmthn = thn+1
    lastday = mlu.DecodeDate(mlu.EncodeDate(nmthn,nmbln,1)-1)
    lastday = lastday[2]
    if tglnum>lastday:
      tglnum=tglnum-lastday
      bln = nmbln
      thn = nmthn
    if tglnum<1:
      lastday = mlu.DecodeDate(mlu.EncodeDate(thn,bln,1)-1)
      thn = lastday[0]
      bln = lastday[1]
      tglnum = lastday[2]
    period = periodGenerate(ptype, tglnum, bln, thn, hari)
    #raise Exception, (period[0],ptype, tglnum, bln, thn, hari)
    s = "select * from period where period_code='%s' and period_type='%s'" % (period[0], ptype)
    res = config.CreateSQL(s).RawResult
    if not res.Eof:
      #raise Exception, 'Ada'
      pass
    else:
      #raise Exception, 'Belum Ada'
      s = "insert into period (period_id, period_code, description, period_type) values (seq_period.nextval, '%s', '%s', '%s')" % (period[0], period[1], ptype)
      #raise Exception, s
      config.ExecSQL(s)
  #return 
  
